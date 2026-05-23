"""
ERAS Systematic Review — Data Extraction App
Streamlit web app for structured data extraction across 7 outcome sheets.
Saves to Excel matching ERAS_DataExtraction_Workbook.xlsx column headers.
"""

import streamlit as st
import pdfplumber
import io
import json
import re
import html
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="ERAS Data Extractor",
    page_icon="🔬",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
  .stTabs [data-baseweb="tab"] { font-size:13px; padding:6px 10px; }
  .record-box {background:#1e2d3d;border-radius:6px;padding:10px 14px;margin:6px 0;font-size:12px;}
  .status-done {color:#00c853;font-weight:600;}
  .status-partial {color:#ffa000;font-weight:600;}
  .status-none {color:#757575;}
  div[data-testid="stNumberInput"] input {font-size:13px;}
  div[data-testid="stTextInput"] input {font-size:13px;}
</style>
""", unsafe_allow_html=True)

# ── Constants ─────────────────────────────────────────────────────────────────
WHO_REGIONS       = ["", "AFR", "AMR", "EMR", "EUR", "SEAR", "WPR"]
INCOME_GROUPS     = ["", "LIC", "LMIC", "UMIC", "HIC"]
STUDY_DESIGNS     = ["", "RCT", "Prospective Cohort", "Retrospective Cohort",
                      "Before-After (Quasi-experimental)", "Cross-sectional",
                      "Case-Control", "Interrupted Time Series", "Other"]
IMPL_TYPES        = ["", "Partial (<50%)", "Moderate (50–70%)", "Comprehensive (>70%)"]
EFFECT_MEASURES   = ["", "MD", "SMD", "RR", "OR", "HR", "RD"]
STATS_METHODS     = ["", "Random-effects (DerSimonian-Laird)", "Fixed-effects",
                      "HKSJ correction", "Peto OR", "Arcsine transformation", "Other"]
YN                = ["", "Y", "N"]
SPECIALTIES       = ["", "Colorectal", "Hepatobiliary", "Upper GI / Oesophagogastric",
                      "Urology", "Gynaecology / Oncology", "Orthopaedics",
                      "Vascular", "Thoracic", "General Surgery", "Mixed", "Other"]
LOS_TYPES         = ["", "Total Hospital", "ICU", "Ward"]
LOS_MEASURES      = ["", "Mean", "Median"]
COMP_SCOPES       = ["", "Overall", "Specific"]
CLAVIEN_GRADES    = ["", "I", "II", "IIIa", "IIIb", "IVa", "IVb", "V", "II+", "IIIa+"]
READ_TIMEFRAMES   = ["", "30-day", "90-day", "1-year", "Other"]
MORT_TYPES        = ["", "In-hospital", "30-day", "90-day", "1-year", "Other"]
COST_PERSP        = ["", "Hospital", "Societal", "Payer", "Patient"]
COST_COMP         = ["", "Direct", "Indirect", "Both"]
ADHERENCE_CATS    = ["", "<50% Partial", "50–70% Moderate", ">70% Comprehensive"]
ADHERENCE_METHODS = ["", "Audit", "Checklist", "Self-report", "Database review", "Other"]
HOSPITAL_LEVELS   = ["", "Primary", "Secondary", "Tertiary", "Academic/University", "Other"]
URBAN_RURAL       = ["", "Urban", "Rural", "Mixed", "Not stated"]
SUBREGS           = ["", "Eastern Africa", "Western Africa", "Northern Africa",
                      "Southern Africa", "Middle Africa", "South Asia", "Southeast Asia",
                      "Eastern Asia", "Western Asia", "Central Asia", "South America",
                      "Central America", "Caribbean", "Northern America",
                      "Northern Europe", "Southern Europe", "Western Europe",
                      "Eastern Europe", "Oceania", "Other"]

OUTCOMES = ["O1_LOS", "O2_COMP", "O3_READ", "O4_MORT", "O5_COST", "ADHERENCE", "GEO_DISPARITY"]
OUTCOME_ICONS = {
    "O1_LOS":       "📏 LOS",
    "O2_COMP":      "🩺 Complications",
    "O3_READ":      "🔄 Readmission",
    "O4_MORT":      "💀 Mortality",
    "O5_COST":      "💰 Cost",
    "ADHERENCE":    "📋 Adherence",
    "GEO_DISPARITY":"🌍 Geography",
}
MULTI_RECORD = {"O1_LOS", "O2_COMP"}   # outcomes supporting multiple rows/study

ERAS_ELEMENTS = [
    ("preop_counsel",    "Preop Counselling / Education"),
    ("preop_nutrition",  "Preop Nutrition Optimisation"),
    ("carb_loading",     "Carbohydrate Loading"),
    ("reduced_fasting",  "Reduced Preop Fasting (Solids <6h, Liquids <2h)"),
    ("preop_analgesia",  "Preop Oral Analgesia / Multimodal Premedication"),
    ("anxiolytic_avoid", "Anxiolytic Avoidance / Minimisation"),
    ("vte_prophylaxis",  "VTE Prophylaxis (Mechanical + Pharmacological)"),
    ("abx_prophylaxis",  "Antibiotic Prophylaxis"),
    ("bowel_prep_omit",  "Bowel Prep Omission"),
    ("ponv_prophylaxis", "PONV Prophylaxis (>=2 agents)"),
    ("std_anaesthesia",  "Standardised Anaesthesia Protocol"),
    ("short_acting_ana", "Short-acting Anaesthetic Agents"),
    ("normothermia",     "Intraoperative Normothermia"),
    ("gdft",             "Goal-directed Fluid Therapy / Restricted Fluids"),
    ("min_invasive",     "Minimally Invasive / Laparoscopic Approach"),
    ("no_ngt",           "No Nasogastric Tube or Early Removal"),
    ("no_drain",         "No / Restricted Abdominal Drains"),
    ("early_cath_rem",   "No / Early Urinary Catheter Removal"),
    ("early_nutrition",  "Early Oral Nutrition (POD 0-1)"),
    ("early_mob",        "Early Mobilisation (POD 0-1)"),
    ("multimodal_analg", "Multimodal Analgesia"),
    ("opioid_sparing",   "Opioid-sparing / Opioid-free Protocol"),
    ("laxatives",        "Laxatives / Prokinetics"),
    ("audit_monitor",    "Audit / Compliance Monitoring"),
    ("other_element1",   "Other Element 1 (specify in Notes)"),
    ("other_element2",   "Other Element 2 (specify in Notes)"),
]

# ── Session state ─────────────────────────────────────────────────────────────
_defaults = {
    "studies":        [],     # [{name, text}]
    "uploaded_names": set(),
    "current_study":  0,
    "extractions":    {},     # {study_idx: {outcome: [record_dicts]}}
    "edit_state":     {},     # {outcome: {field: value}}   ← current form draft
    "edit_record":    {},     # {outcome: record_idx or None (=new)}
    "excel_bytes":    None,
    "search_term":    "",
}
for _k, _v in _defaults.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v
ss = st.session_state


# ── PDF helpers ───────────────────────────────────────────────────────────────
def extract_pdf(file_bytes: bytes) -> str:
    try:
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            return "\n\n".join(
                p.extract_text() or "" for p in pdf.pages
            ).strip()
    except Exception as e:
        return f"[Extraction error: {e}]"


# ── Text viewer with highlighting ─────────────────────────────────────────────
_HLIGHT = {
    "LOS / length of stay": "#FFE082",
    "complication":          "#FFAB91",
    "readmission":           "#CE93D8",
    "mortalit":              "#EF9A9A",
    "cost / cost-effective": "#80DEEA",
    "adherence / compliance":"#A5D6A7",
    "ERAS / enhanced recov": "#FFF176",
}
_HLIGHT_PATTERNS = {
    re.compile(r"\b(LOS|length of stay|hospital stay|discharge)\b", re.I): "#FFE082",
    re.compile(r"\b(complication|morbidity|adverse|SSI|leak|ileus|infection)\b", re.I): "#FFAB91",
    re.compile(r"\b(readmission|re-admission|readmit)\b", re.I): "#CE93D8",
    re.compile(r"\b(mortalit|death|died|fatal)\b", re.I): "#EF9A9A",
    re.compile(r"\b(cost|expenditure|economic|PPP|ICER|saving)\b", re.I): "#80DEEA",
    re.compile(r"\b(adherence|compliance|protocol element)\b", re.I): "#A5D6A7",
    re.compile(r"\b(ERAS|enhanced recovery|fast.track|perioperative pathway)\b", re.I): "#FFF176",
}

def build_viewer(text: str, search: str) -> str:
    escaped = html.escape(text)
    # Map original→escaped positions (html.escape can expand &, <, >)
    orig_to_esc = []
    ei = 0
    for ch in text:
        orig_to_esc.append(ei)
        ec = html.escape(ch)
        ei += len(ec)
    orig_to_esc.append(ei)

    marks: list[tuple[int,int,str]] = []  # (start_esc, end_esc, color)
    for pat, col in _HLIGHT_PATTERNS.items():
        for m in pat.finditer(text):
            marks.append((orig_to_esc[m.start()], orig_to_esc[m.end()], col))
    if search.strip():
        spat = re.compile(re.escape(search.strip()), re.I)
        for m in spat.finditer(text):
            marks.append((orig_to_esc[m.start()], orig_to_esc[m.end()], "#FF6F00"))

    marks.sort(key=lambda x: x[0])
    parts = []
    pos = 0
    for s, e, col in marks:
        if s < pos:
            continue
        parts.append(escaped[pos:s])
        parts.append(f'<mark style="background:{col};color:#111;padding:0 2px;border-radius:2px">')
        parts.append(escaped[s:e])
        parts.append("</mark>")
        pos = e
    parts.append(escaped[pos:])
    body = "".join(parts)
    return (
        '<div style="font-family:Consolas,\'Courier New\',monospace;font-size:12px;'
        'line-height:1.7;white-space:pre-wrap;word-wrap:break-word;'
        'background:#FFFFFF;color:#1a1a1a;padding:18px 22px;border-radius:8px;'
        'border:1px solid #dee2e6;height:680px;overflow-y:auto;">' + body + "</div>"
    )


# ── Extraction helpers ────────────────────────────────────────────────────────
def get_records(study_idx: int, outcome: str) -> list:
    return ss.extractions.setdefault(study_idx, {}).setdefault(outcome, [])

def save_record(study_idx: int, outcome: str, rec: dict, rec_idx):
    recs = get_records(study_idx, outcome)
    if rec_idx is None:
        recs.append(rec)
    else:
        recs[rec_idx] = rec

def delete_record(study_idx: int, outcome: str, rec_idx: int):
    recs = get_records(study_idx, outcome)
    if 0 <= rec_idx < len(recs):
        recs.pop(rec_idx)

def init_edit(outcome: str, defaults: dict = None):
    if outcome not in ss.edit_state:
        ss.edit_state[outcome] = defaults or {}
    if outcome not in ss.edit_record:
        ss.edit_record[outcome] = None


def _apply_extracted(outcome: str, study_idx: int, data: dict):
    """
    Write extracted values into BOTH edit_state AND Streamlit's widget
    session-state cache.  Without the latter, widgets keep showing their
    previously-rendered (cached) values even after edit_state is updated.
    """
    ss.edit_state[outcome] = data
    ss.edit_record[outcome] = None

    pfx = f"{outcome}_{study_idx}"   # widget key prefix used by every form

    def _set(suffix, val):
        if val is not None and val != "":
            ss[f"{pfx}{suffix}"] = val

    # ── Common identification fields ──────────────────────────────────────────
    _set("_sid",   str(data.get("study_id", "")))
    _set("_auth",  str(data.get("first_author", "")))
    if data.get("year"):       _set("_year", int(data["year"]))
    _set("_cntry", str(data.get("country", "")))
    _set("_who",   str(data.get("who_region", "")))
    _set("_inc",   str(data.get("income_group", "")))
    _set("_spec",  str(data.get("surgical_specialty", "")))
    _set("_proc",  str(data.get("specific_procedure", "")))
    _set("_des",   str(data.get("study_design", "")))
    if data.get("n_total"):    _set("_nt", int(data["n_total"]))
    if data.get("n_eras"):     _set("_ne", int(data["n_eras"]))
    if data.get("n_control"):  _set("_nc", int(data["n_control"]))

    # ── Meta-analysis fields ──────────────────────────────────────────────────
    if data.get("p_value"):    _set("_pv",  float(data["p_value"]))
    if data.get("effect_size"):_set("_es",  float(data["effect_size"]))

    # ── O1_LOS ───────────────────────────────────────────────────────────────
    if outcome == "O1_LOS":
        _set("_ltype", str(data.get("los_type", "")))
        _set("_lmeas", str(data.get("los_measure", "")))
        if data.get("los_eras"):        _set("_le",   float(data["los_eras"]))
        if data.get("los_eras_sd"):     _set("_les",  float(data["los_eras_sd"]))
        if data.get("eras_iqr_lower"):  _set("_leil", float(data["eras_iqr_lower"]))
        if data.get("eras_iqr_upper"):  _set("_leiu", float(data["eras_iqr_upper"]))
        if data.get("los_control"):     _set("_lc",   float(data["los_control"]))
        if data.get("los_ctrl_sd"):     _set("_lcs",  float(data["los_ctrl_sd"]))
        if data.get("ctrl_iqr_lower"):  _set("_lcil", float(data["ctrl_iqr_lower"]))
        if data.get("ctrl_iqr_upper"):  _set("_lciu", float(data["ctrl_iqr_upper"]))
        if data.get("mean_diff"):       _set("_md",   float(data["mean_diff"]))
        if data.get("pct_reduction"):   _set("_pr",   float(data["pct_reduction"]))

    # ── O2_COMP ──────────────────────────────────────────────────────────────
    elif outcome == "O2_COMP":
        _set("_cs",   str(data.get("complication_scope", "")))
        _set("_ct",   str(data.get("complication_type", "")))
        if data.get("events_eras") is not None:    _set("_ee", int(data["events_eras"]))
        if data.get("events_control") is not None: _set("_ec", int(data["events_control"]))
        if data.get("rate_eras"):    _set("_re",   float(data["rate_eras"]))
        if data.get("rate_ctrl"):    _set("_rc",   float(data["rate_ctrl"]))
        _set("_ssi",  str(data.get("ssi_reported",  "")))
        _set("_al",   str(data.get("al_reported",   "")))
        _set("_poi",  str(data.get("poi_reported",  "")))
        _set("_pulm", str(data.get("pulm_reported", "")))
        _set("_urin", str(data.get("urin_reported", "")))

    # ── O3_READ ──────────────────────────────────────────────────────────────
    elif outcome == "O3_READ":
        _set("_tf",  str(data.get("readmission_timeframe", "")))
        if data.get("events_eras") is not None:    _set("_ee", int(data["events_eras"]))
        if data.get("events_control") is not None: _set("_ec", int(data["events_control"]))

    # ── O4_MORT ──────────────────────────────────────────────────────────────
    elif outcome == "O4_MORT":
        _set("_mtype", str(data.get("mortality_type", "")))
        if data.get("events_eras") is not None:    _set("_ee", int(data["events_eras"]))
        if data.get("events_control") is not None: _set("_ec", int(data["events_control"]))

    # ── ADHERENCE ────────────────────────────────────────────────────────────
    elif outcome == "ADHERENCE":
        if data.get("overall_adherence_rate"): _set("_oar", float(data["overall_adherence_rate"]))
        _set("_acat", str(data.get("adherence_category", "")))
        for el_key, _ in ERAS_ELEMENTS:
            if data.get(el_key) is not None:
                _set(f"_{el_key}", float(data[el_key]))

    # ── GEO_DISPARITY ────────────────────────────────────────────────────────
    elif outcome == "GEO_DISPARITY":
        _set("_subreg", str(data.get("subregion", "")))
        _set("_ur",     str(data.get("urban_rural", "")))
        _set("_hosp",   str(data.get("hospital_level", "")))
        if data.get("overall_adherence_rate"): _set("_ar", float(data["overall_adherence_rate"]))


# ── Smart Extraction ─────────────────────────────────────────────────────────
# Flexible plus-minus: handles U+00B1, literal ±, or +/-
_PM = r'(?:\xb1|±|\+\s*/\s*-|\+\-)\s*'

_COUNTRY_DB = {
    'india': ('India', 'SEAR', 'LMIC'),
    'ethiopia': ('Ethiopia', 'AFR', 'LIC'),
    'uganda': ('Uganda', 'AFR', 'LIC'),
    'nigeria': ('Nigeria', 'AFR', 'LMIC'),
    'ghana': ('Ghana', 'AFR', 'LMIC'),
    'kenya': ('Kenya', 'AFR', 'LMIC'),
    'tanzania': ('Tanzania', 'AFR', 'LIC'),
    'malawi': ('Malawi', 'AFR', 'LIC'),
    'zambia': ('Zambia', 'AFR', 'LMIC'),
    'rwanda': ('Rwanda', 'AFR', 'LIC'),
    'south africa': ('South Africa', 'AFR', 'UMIC'),
    'egypt': ('Egypt', 'EMR', 'LMIC'),
    'iran': ('Iran', 'EMR', 'UMIC'),
    'jordan': ('Jordan', 'EMR', 'UMIC'),
    'pakistan': ('Pakistan', 'EMR', 'LMIC'),
    'bangladesh': ('Bangladesh', 'SEAR', 'LMIC'),
    'nepal': ('Nepal', 'SEAR', 'LIC'),
    'sri lanka': ('Sri Lanka', 'SEAR', 'LMIC'),
    'thailand': ('Thailand', 'SEAR', 'UMIC'),
    'indonesia': ('Indonesia', 'SEAR', 'LMIC'),
    'china': ('China', 'WPR', 'UMIC'),
    'malaysia': ('Malaysia', 'WPR', 'UMIC'),
    'philippines': ('Philippines', 'WPR', 'LMIC'),
    'vietnam': ('Vietnam', 'WPR', 'LMIC'),
    'netherlands': ('Netherlands', 'EUR', 'HIC'),
    'united kingdom': ('United Kingdom', 'EUR', 'HIC'),
    'germany': ('Germany', 'EUR', 'HIC'),
    'france': ('France', 'EUR', 'HIC'),
    'sweden': ('Sweden', 'EUR', 'HIC'),
    'norway': ('Norway', 'EUR', 'HIC'),
    'denmark': ('Denmark', 'EUR', 'HIC'),
    'turkey': ('Turkey', 'EUR', 'UMIC'),
    'usa': ('United States', 'AMR', 'HIC'),
    'united states': ('United States', 'AMR', 'HIC'),
    'canada': ('Canada', 'AMR', 'HIC'),
    'brazil': ('Brazil', 'AMR', 'UMIC'),
    'australia': ('Australia', 'WPR', 'HIC'),
    'new zealand': ('New Zealand', 'WPR', 'HIC'),
    'uk': ('United Kingdom', 'EUR', 'HIC'),
}

def smart_extract_common(text: str) -> dict:
    from collections import Counter
    d = {}
    # ── Study design ──────────────────────────────────────────────────────────
    if re.search(r'\b(randomized|randomised)\s+controlled\s+trial\b|\bRCT\b', text, re.I):
        d['study_design'] = 'RCT'
    elif re.search(r'\bprospective\s+cohort\b', text, re.I):
        d['study_design'] = 'Prospective Cohort'
    elif re.search(r'\bretrospective\b', text, re.I):
        d['study_design'] = 'Retrospective Cohort'
    elif re.search(r'\bbefore[\s\-]after\b|\bpre[\s\-]post\b|\bquasi[\s\-]experimental\b', text, re.I):
        d['study_design'] = 'Before-After (Quasi-experimental)'
    elif re.search(r'\binterrupted\s+time\s+series\b', text, re.I):
        d['study_design'] = 'Interrupted Time Series'

    # ── Country (longest key first to avoid 'uk' matching before 'united kingdom') ──
    tl = text.lower()
    for kw in sorted(_COUNTRY_DB.keys(), key=len, reverse=True):
        if re.search(r'\b' + re.escape(kw) + r'\b', tl):
            d['country'], d['who_region'], d['income_group'] = _COUNTRY_DB[kw]
            break

    # ── N values ──────────────────────────────────────────────────────────────
    # Total: allow words between "participants" and assignment verb
    m = re.search(
        r'(\d{2,4})\s+(?:participants|patients|women|subjects|adults)'
        r'[^.]{0,150}?(?:randomly\s+)?(?:assigned|enrolled|randomized|randomised|recruited)',
        text, re.I | re.S)
    if m: d['n_total'] = int(m.group(1))

    # ERAS group N
    for pat in [
        r'(\d+)\s+(?:were\s+)?(?:allocated|assigned|randomized|randomised)\s+to\s+(?:the\s+)?ERAS',
        r'ERAS\s+group\s*\(\s*[nN]\s*=\s*(\d+)',
        r'ERAS\s+\(\s*[nN]\s*=\s*(\d+)',
        r'ERAS\s+group[^\n]*?n\s*=\s*(\d+)',
    ]:
        m = re.search(pat, text, re.I)
        if m: d['n_eras'] = int(m.group(1)); break

    # Control group N — includes "and 50 to the control group" pattern
    for pat in [
        r'(\d+)\s+(?:were\s+)?(?:allocated|assigned|randomized|randomised)\s+to\s+(?:the\s+)?control',
        r'and\s+(\d+)\s+to\s+(?:the\s+)?control',
        r'(\d+)\s+to\s+(?:the\s+)?control\s+group',
        r'control\s+group\s*\(\s*[nN]\s*=\s*(\d+)',
        r'control\s+\(\s*[nN]\s*=\s*(\d+)',
    ]:
        m = re.search(pat, text, re.I)
        if m: d['n_control'] = int(m.group(1)); break

    if 'n_eras' in d and 'n_control' in d and 'n_total' not in d:
        d['n_total'] = d['n_eras'] + d['n_control']

    # ── Year — prefer journal citation year e.g. "2025;17:" over reference years ──
    m = re.search(r'\b(20[1-2]\d)\s*;\s*\d+', text)   # "2025;17:S1849"
    if not m:
        m = re.search(r'(?:Published|Accepted|Submitted)[^\n]{0,20}(20[1-2]\d)', text, re.I)
    if not m:
        years = re.findall(r'\b(20[1-2]\d)\b', text)
        if years: d['year'] = int(Counter(years).most_common(1)[0][0])
    if m: d['year'] = int(m.group(1))

    # ── First author — prefer "G. D. Abilashini" initials pattern ─────────────
    am = re.search(r'[A-Z]\.\s+[A-Z]\.\s+([A-Z][a-z]+)', text)        # G. D. Lastname
    if not am:
        am = re.search(r'\b([A-Z][a-z]{3,})\s+[A-Z]{1,2},', text)     # Lastname GD,
    if not am:
        am = re.search(r'(?:^|\n)([A-Z][a-z]{3,}),', text)             # Lastname, First…
    if am:
        d['first_author'] = am.group(1)
        if d.get('year'):
            abbr = re.sub(r'[^A-Za-z]', '', d['first_author'])[:2].upper()
            d['study_id'] = f"{abbr}{d['year']}01"

    # ── Surgical specialty ────────────────────────────────────────────────────
    for pat, spec in [
        (r'\b(colorectal|colon\s+cancer|rectal\s+cancer|colectomy|sigmoid)\b', 'Colorectal'),
        (r'\b(hepatobiliary|hepatic|liver\s+resection|pancreatectomy|cholecystectomy)\b', 'Hepatobiliary'),
        (r'\b(oesophag|esophag|gastrectomy|gastric\s+cancer|upper\s+GI)\b', 'Upper GI / Oesophagogastric'),
        (r'\b(cystectomy|nephrectomy|prostatectomy|urolog)\b', 'Urology'),
        (r'\b(gynae|gynecol|obstetric|cesarean|caesarean|hysterectomy|ovarian)\b', 'Gynaecology / Oncology'),
        (r'\b(arthroplasty|hip\s+replacement|knee\s+replacement|orthop)\b', 'Orthopaedics'),
        (r'\b(aortic|aorta|vascular|endarterectomy)\b', 'Vascular'),
        (r'\b(thorac|lung\s+resection|lobectomy|pneumonectomy)\b', 'Thoracic'),
    ]:
        if re.search(pat, text, re.I): d['surgical_specialty'] = spec; break

    # ── Specific procedure (short snippet) ────────────────────────────────────
    for pat in [
        r'(?:emergency|elective)\s+[a-z\s]{3,35}?(?:delivery|section|surgery|resection|repair)',
        r'(?:laparoscopic|open)\s+[a-z\s]{3,35}?(?:resection|repair|colectomy|gastrectomy)',
    ]:
        pm = re.search(pat, text, re.I)
        if pm: d['specific_procedure'] = pm.group(0)[:60].strip(); break

    return d


def smart_extract_los(text: str) -> dict:
    d = {}
    # ── Table row: "Length of hospital stay  3.2±1.1  5.4±1.3" ───────────────
    # Use [\s\S] so it works whether values are on same line or next cell
    _val = r'(\d+\.\d+)\s*' + _PM + r'(\d+\.\d+)'
    m = re.search(
        r'[Ll]ength\s+of\s+(?:hospital\s+)?stay[\s\S]{0,80}?' + _val +
        r'[\s\S]{0,80}?' + _val,
        text)
    if m:
        d.update(los_type='Total Hospital', los_measure='Mean',
                 los_eras=float(m.group(1)), los_eras_sd=float(m.group(2)),
                 los_control=float(m.group(3)), los_ctrl_sd=float(m.group(4)))

    # ── Prose: "shorter hospital stay (3.2 ± 1.1 days)…control group (5.4 ± 1.3 days)" ──
    if 'los_eras' not in d:
        m2 = re.search(
            r'(?:shorter|longer)\s+hospital\s+stay\s*\((\d+\.\d+)\s*' + _PM +
            r'(\d+\.\d+)[^)]*\)[\s\S]{0,200}?(?:control|comparison)[^\(]{0,60}\((\d+\.\d+)\s*' + _PM + r'(\d+\.\d+)',
            text, re.I | re.S)
        if m2:
            d.update(los_type='Total Hospital', los_measure='Mean',
                     los_eras=float(m2.group(1)), los_eras_sd=float(m2.group(2)),
                     los_control=float(m2.group(3)), los_ctrl_sd=float(m2.group(4)))

    # ── Median IQR: "X.X (X.X–X.X)" ──────────────────────────────────────────
    if 'los_eras' not in d:
        m3 = re.search(
            r'[Ll]ength\s+of\s+(?:hospital\s+)?stay[\s\S]{0,100}'
            r'(\d+\.?\d*)\s*\((\d+\.?\d*)\s*[–\-]\s*(\d+\.?\d*)\)[\s\S]{0,100}'
            r'(\d+\.?\d*)\s*\((\d+\.?\d*)\s*[–\-]\s*(\d+\.?\d*)\)',
            text)
        if m3:
            d.update(los_type='Total Hospital', los_measure='Median',
                     los_eras=float(m3.group(1)), eras_iqr_lower=float(m3.group(2)),
                     eras_iqr_upper=float(m3.group(3)),
                     los_control=float(m3.group(4)), ctrl_iqr_lower=float(m3.group(5)),
                     ctrl_iqr_upper=float(m3.group(6)))

    # ── Derived fields ────────────────────────────────────────────────────────
    if 'los_eras' in d and 'los_control' in d:
        diff = round(d['los_control'] - d['los_eras'], 3)
        d['mean_diff'] = diff
        if d['los_control'] > 0:
            d['pct_reduction'] = round(diff / d['los_control'] * 100, 1)

    # ── p-value near LOS (handles "p<0.001", "<0.001", "0.001*") ─────────────
    pm = re.search(
        r'[Ll]ength\s+of\s+(?:hospital\s+)?stay[\s\S]{0,300}?[pP]\s*[<=>]\s*([0-9.]+)',
        text, re.S)
    if not pm:
        # table row ends with <0.001* — pick up raw value after SD pair
        pm = re.search(
            r'[Ll]ength\s+of\s+(?:hospital\s+)?stay[\s\S]{0,120}?\d+\.\d+\s*[<]\s*([0-9.]+)',
            text, re.S)
    if pm:
        try: d['p_value'] = float(pm.group(1))
        except: pass
    return d


def smart_extract_comp(text: str, n_eras: int = 0, n_ctrl: int = 0) -> dict:
    d = {}
    # Table row may have "Incidence of postoperative  2  8  0.03*\ncomplications"
    # (PDF column flow splits label across lines) — use re.S
    m = re.search(
        r'postoperative\s+(?:complications?\s+)?(\d+)\s+(\d+)\s+([0-9.<>*]+)',
        text, re.I | re.S)
    if not m:
        m = re.search(
            r'[Ii]ncidence[\s\S]{0,60}?(\d+)\s+(\d+)\s+([0-9.<>*]+)',
            text, re.I | re.S)
    if m:
        ev_e, ev_c = int(m.group(1)), int(m.group(2))
        d.update(complication_scope='Overall',
                 complication_type='Overall postoperative complications',
                 events_eras=ev_e, events_control=ev_c)
        if n_eras > 0: d['rate_eras'] = round(ev_e / n_eras * 100, 2)
        if n_ctrl > 0: d['rate_ctrl'] = round(ev_c / n_ctrl * 100, 2)
        try: d['p_value'] = float(re.sub(r'[<>*]', '', m.group(3)))
        except: pass
    # Specific complication types
    if re.search(r'\bSSI\b|surgical\s+site\s+infection', text, re.I):  d['ssi_reported']  = 'Y'
    if re.search(r'anastomotic\s+leak', text, re.I):                    d['al_reported']   = 'Y'
    if re.search(r'\bileus\b|\bPOI\b|paralytic\s+ileus', text, re.I):  d['poi_reported']  = 'Y'
    if re.search(r'pulmonar|pneumonia', text, re.I):                    d['pulm_reported'] = 'Y'
    if re.search(r'\bUTI\b|urinary\s+(?:tract\s+)?infection', text, re.I): d['urin_reported'] = 'Y'
    return d


def smart_extract_read(text: str) -> dict:
    d = {}
    m = re.search(r'readmission[\s\S]{0,250}?(\d+)\s+(\d+)', text, re.I | re.S)
    if m: d.update(readmission_timeframe='30-day',
                   events_eras=int(m.group(1)), events_control=int(m.group(2)))
    return d


def smart_extract_mort(text: str) -> dict:
    d = {}
    m = re.search(r'(?:mortalit|death|died)[\s\S]{0,200}?(\d+)\s+(\d+)', text, re.I | re.S)
    if m: d.update(mortality_type='In-hospital',
                   events_eras=int(m.group(1)), events_control=int(m.group(2)))
    return d


def smart_extract_adherence(text: str) -> dict:
    d = {}
    am = re.search(r'(?:overall\s+)?adherence\s+(?:rate\s+)?(?:of\s+)?([0-9.]+)\s*%', text, re.I)
    if am:
        ar = float(am.group(1))
        d['overall_adherence_rate'] = ar
        if ar >= 70:   d['adherence_category'] = '>70% Comprehensive'
        elif ar >= 50: d['adherence_category'] = '50–70% Moderate'
        else:          d['adherence_category'] = '<50% Partial'
    for key, pat in [
        ('preop_counsel',    r'preop(?:erative)?\s+(?:counsel|education|information)'),
        ('carb_loading',     r'carbohydrate\s+loading|carb\s+load'),
        ('reduced_fasting',  r'reduced\s+fasting|clear\s+fluid'),
        ('vte_prophylaxis',  r'VTE|thromboprophylaxis|compression\s+stock'),
        ('abx_prophylaxis',  r'antibiotic\s+prophylaxis'),
        ('ponv_prophylaxis', r'PONV|anti.?emetic'),
        ('normothermia',     r'normothermia|temperature\s+management|warming'),
        ('gdft',             r'goal.directed|fluid\s+management|fluid\s+restriction'),
        ('min_invasive',     r'laparoscop|minimally\s+invasive'),
        ('no_ngt',           r'nasogastric\s+tube|\bNGT\b'),
        ('early_nutrition',  r'early\s+(?:oral\s+)?(?:nutrition|feeding|intake)'),
        ('early_mob',        r'early\s+mobiliz|early\s+ambul'),
        ('multimodal_analg', r'multimodal\s+analgesia'),
        ('opioid_sparing',   r'opioid.spar|opioid.free|non.opioid'),
        ('audit_monitor',    r'\baudit\b|compliance\s+monitor'),
    ]:
        if re.search(pat, text, re.I): d[key] = 1
    return d


def smart_extract_all(text: str, outcome: str) -> dict:
    d = smart_extract_common(text)
    ne, nc = d.get('n_eras', 0), d.get('n_control', 0)
    if outcome == "O1_LOS":      d.update(smart_extract_los(text))
    elif outcome == "O2_COMP":   d.update(smart_extract_comp(text, ne, nc))
    elif outcome == "O3_READ":   d.update(smart_extract_read(text))
    elif outcome == "O4_MORT":   d.update(smart_extract_mort(text))
    elif outcome == "ADHERENCE": d.update(smart_extract_adherence(text))
    return d


# ── Common identification fields ──────────────────────────────────────────────
def common_id_form(key_prefix: str, defaults: dict) -> dict:
    """Render the study identification block; returns dict of values."""
    d = defaults
    with st.expander("📌 Study Identification (common to all outcomes)", expanded=True):
        c1, c2, c3 = st.columns(3)
        study_id    = c1.text_input("Study ID *",       d.get("study_id",""),    key=f"{key_prefix}_sid",
                                    help="Format: [2 letters][Year][Seq] e.g. SM202301")
        first_author= c2.text_input("First Author *",   d.get("first_author",""),key=f"{key_prefix}_auth")
        year        = c3.number_input("Year *",          min_value=1990, max_value=2030,
                                      value=int(d.get("year", 2020)),         key=f"{key_prefix}_year")
        c4, c5, c6 = st.columns(3)
        country     = c4.text_input("Country",           d.get("country",""),    key=f"{key_prefix}_cntry")
        who_region  = c5.selectbox("WHO Region",         WHO_REGIONS,
                                   index=WHO_REGIONS.index(d.get("who_region","")) if d.get("who_region","") in WHO_REGIONS else 0,
                                   key=f"{key_prefix}_who")
        income_grp  = c6.selectbox("Income Group",       INCOME_GROUPS,
                                   index=INCOME_GROUPS.index(d.get("income_group","")) if d.get("income_group","") in INCOME_GROUPS else 0,
                                   key=f"{key_prefix}_inc")
        c7, c8 = st.columns(2)
        specialty   = c7.selectbox("Surgical Specialty", SPECIALTIES,
                                   index=SPECIALTIES.index(d.get("surgical_specialty","")) if d.get("surgical_specialty","") in SPECIALTIES else 0,
                                   key=f"{key_prefix}_spec")
        procedure   = c8.text_input("Specific Procedure",d.get("specific_procedure",""), key=f"{key_prefix}_proc")
        c9, c10, c11, c12 = st.columns(4)
        study_design= c9.selectbox("Study Design",       STUDY_DESIGNS,
                                   index=STUDY_DESIGNS.index(d.get("study_design","")) if d.get("study_design","") in STUDY_DESIGNS else 0,
                                   key=f"{key_prefix}_des")
        n_total     = c10.number_input("N Total",         min_value=0, value=int(d.get("n_total",0)),     key=f"{key_prefix}_nt")
        n_eras      = c11.number_input("N ERAS",          min_value=0, value=int(d.get("n_eras",0)),      key=f"{key_prefix}_ne")
        n_control   = c12.number_input("N Control",       min_value=0, value=int(d.get("n_control",0)),   key=f"{key_prefix}_nc")
        c13, c14 = st.columns(2)
        adh_rate    = c13.number_input("Adherence Rate (%)", min_value=0.0, max_value=100.0, step=0.1,
                                       value=float(d.get("adherence_rate",0.0)),                         key=f"{key_prefix}_adhr")
        impl_type   = c14.selectbox("Implementation Type", IMPL_TYPES,
                                    index=IMPL_TYPES.index(d.get("implementation_type","")) if d.get("implementation_type","") in IMPL_TYPES else 0,
                                    key=f"{key_prefix}_impl")
    return dict(study_id=study_id, first_author=first_author, year=int(year),
                country=country, who_region=who_region, income_group=income_grp,
                surgical_specialty=specialty, specific_procedure=procedure,
                study_design=study_design, n_total=int(n_total),
                n_eras=int(n_eras), n_control=int(n_control),
                adherence_rate=adh_rate, implementation_type=impl_type)


def meta_fields(key_prefix: str, d: dict) -> dict:
    """Render shared meta-analysis fields."""
    st.markdown("**Meta-Analysis Data**")
    c1, c2, c3, c4 = st.columns(4)
    effect_m = c1.selectbox("Effect Measure", EFFECT_MEASURES,
                             index=EFFECT_MEASURES.index(d.get("effect_measure","")) if d.get("effect_measure","") in EFFECT_MEASURES else 0,
                             key=f"{key_prefix}_em")
    effect_s = c2.number_input("Effect Size",  value=float(d.get("effect_size") or 0), step=0.001, format="%.4f", key=f"{key_prefix}_es")
    ci_lo    = c3.number_input("95% CI Lower", value=float(d.get("ci_lower") or 0),    step=0.001, format="%.4f", key=f"{key_prefix}_cil")
    ci_hi    = c4.number_input("95% CI Upper", value=float(d.get("ci_upper") or 0),    step=0.001, format="%.4f", key=f"{key_prefix}_ciu")
    c5, c6, c7, c8 = st.columns(4)
    p_val    = c5.number_input("p-value",      value=float(d.get("p_value") or 0),      step=0.001, format="%.4f", key=f"{key_prefix}_pv")
    i2       = c6.number_input("I² (%)",       value=float(d.get("i2") or 0),           step=0.1,   format="%.1f", key=f"{key_prefix}_i2")
    tau2     = c7.number_input("Tau²",         value=float(d.get("tau2") or 0),          step=0.001, format="%.4f", key=f"{key_prefix}_tau")
    stat_m   = c8.selectbox("Statistical Method", STATS_METHODS,
                              index=STATS_METHODS.index(d.get("stat_method","")) if d.get("stat_method","") in STATS_METHODS else 0,
                              key=f"{key_prefix}_sm")
    c9, c10, c11, c12 = st.columns(4)
    imput    = c9.selectbox("Imputation Used", YN,
                             index=YN.index(d.get("imputation_used","")) if d.get("imputation_used","") in YN else 0,
                             key=f"{key_prefix}_imp")
    extr_by  = c10.text_input("Extracted By",  d.get("extracted_by",""),  key=f"{key_prefix}_exby")
    dt_str   = d.get("date_extracted", str(date.today()))
    try:
        dt_val = datetime.strptime(dt_str, "%Y-%m-%d").date() if dt_str else date.today()
    except Exception:
        dt_val = date.today()
    dt_ext   = c11.date_input("Date Extracted", dt_val, key=f"{key_prefix}_dt")
    verif_by = c12.text_input("Verified By",    d.get("verified_by",""),  key=f"{key_prefix}_vby")
    notes    = st.text_area("Notes", d.get("notes",""), height=80, key=f"{key_prefix}_notes")
    return dict(effect_measure=effect_m, effect_size=effect_s, ci_lower=ci_lo,
                ci_upper=ci_hi, p_value=p_val, i2=i2, tau2=tau2, stat_method=stat_m,
                imputation_used=imput, extracted_by=extr_by,
                date_extracted=str(dt_ext), verified_by=verif_by, notes=notes)


def _fmt_mean_sd(mean, sd):
    """Return 'X.XX ± Y.YY' or just 'X.XX' if SD is blank/zero."""
    m = f"{float(mean):.2f}" if mean not in (None, "", 0, 0.0) else ""
    s = f"{float(sd):.2f}"   if sd   not in (None, "", 0, 0.0) else ""
    if m and s:  return f"{m} ± {s}"
    return m or "—"

def record_summary(rec: dict, outcome: str) -> str:
    """One-line summary of a saved record."""
    parts = []
    if outcome == "O1_LOS":
        eras_str = _fmt_mean_sd(rec.get('los_eras'), rec.get('los_eras_sd'))
        ctrl_str = _fmt_mean_sd(rec.get('los_control'), rec.get('los_ctrl_sd'))
        parts = [rec.get("los_type",""), rec.get("los_measure",""),
                 f"ERAS: {eras_str} d", f"Ctrl: {ctrl_str} d"]
    elif outcome == "O2_COMP":
        parts = [rec.get("complication_type","Overall"), rec.get("complication_scope",""),
                 f"ERAS: {rec.get('events_eras','')} n", f"Ctrl: {rec.get('events_control','')} n",
                 f"Rate ERAS: {rec.get('rate_eras','')}%"]
    elif outcome == "O3_READ":
        parts = [rec.get("readmission_timeframe",""), f"ERAS: {rec.get('readmit_rate_eras','')}%"]
    elif outcome == "O4_MORT":
        parts = [rec.get("mortality_type",""), f"ERAS: {rec.get('mort_rate_eras','')}%"]
    elif outcome == "O5_COST":
        parts = [rec.get("cost_perspective",""), rec.get("original_currency",""),
                 f"Diff: {rec.get('cost_diff_orig','')}"]
    elif outcome == "ADHERENCE":
        parts = [f"Adh: {rec.get('overall_adherence_rate','')}%",
                 rec.get("adherence_category","")]
    elif outcome == "GEO_DISPARITY":
        parts = [rec.get("country",""), rec.get("who_region",""),
                 f"Adh:{rec.get('overall_adherence_rate','')}%"]
    return "  |  ".join(str(p) for p in parts if p)


def records_panel(study_idx: int, outcome: str):
    """Show existing records with edit/delete buttons."""
    recs = get_records(study_idx, outcome)
    if not recs:
        st.info("No records saved yet for this outcome.")
        return
    st.markdown(f"**Saved records ({len(recs)})**")
    for i, rec in enumerate(recs):
        c1, c2, c3 = st.columns([6, 1, 1])
        c1.markdown(f'<div class="record-box">#{i+1} &nbsp; {record_summary(rec, outcome)}</div>',
                    unsafe_allow_html=True)
        if c2.button("✏️", key=f"edit_{outcome}_{study_idx}_{i}", help="Edit"):
            ss.edit_state[outcome] = dict(rec)
            ss.edit_record[outcome] = i
            st.rerun()
        if c3.button("🗑️", key=f"del_{outcome}_{study_idx}_{i}", help="Delete"):
            delete_record(study_idx, outcome, i)
            if ss.edit_record.get(outcome) == i:
                ss.edit_record[outcome] = None
                ss.edit_state[outcome] = {}
            st.rerun()


# ── O1_LOS form ──────────────────────────────────────────────────────────────
def form_o1(study_idx: int):
    outcome = "O1_LOS"
    init_edit(outcome)
    records_panel(study_idx, outcome)

    is_new = ss.edit_record.get(outcome) is None
    st.markdown(f"### {'➕ New Record' if is_new else f'✏️ Editing Record #{ss.edit_record[outcome]+1}'}")
    d = ss.edit_state.get(outcome, {})
    id_data = common_id_form(f"{outcome}_{study_idx}", d)

    st.markdown("**Length of Stay Data**")
    c1, c2 = st.columns(2)
    los_type    = c1.selectbox("LOS Type", LOS_TYPES,
                                index=LOS_TYPES.index(d.get("los_type","")) if d.get("los_type","") in LOS_TYPES else 0,
                                key=f"{outcome}_{study_idx}_ltype",
                                help="Total Hospital LOS, ICU LOS, or Ward LOS")
    los_measure = c2.selectbox("LOS Measure", LOS_MEASURES,
                                index=LOS_MEASURES.index(d.get("los_measure","")) if d.get("los_measure","") in LOS_MEASURES else 0,
                                key=f"{outcome}_{study_idx}_lmeas",
                                help="Whether reported as Mean or Median")

    # ERAS group — mean/median and dispersion in clearly labelled separate columns
    st.markdown("<small style='color:#7EC8E3'>🔵 ERAS group</small>", unsafe_allow_html=True)
    c3, c4, c5, c6 = st.columns(4)
    los_eras    = c3.number_input("ERAS — Mean / Median (days)", value=float(d.get("los_eras") or 0),       step=0.1, format="%.2f", key=f"{outcome}_{study_idx}_le")
    los_eras_sd = c4.number_input("ERAS — SD",                   value=float(d.get("los_eras_sd") or 0),    step=0.1, format="%.2f", key=f"{outcome}_{study_idx}_les",  help="SD if Mean reported")
    eras_iqr_lo = c5.number_input("ERAS — IQR Lower",            value=float(d.get("eras_iqr_lower") or 0), step=0.1, format="%.2f", key=f"{outcome}_{study_idx}_leil", help="IQR lower if Median reported")
    eras_iqr_hi = c6.number_input("ERAS — IQR Upper",            value=float(d.get("eras_iqr_upper") or 0), step=0.1, format="%.2f", key=f"{outcome}_{study_idx}_leiu", help="IQR upper if Median reported")

    # Control group — mean/median and dispersion
    st.markdown("<small style='color:#F4A261'>🟠 Control group</small>", unsafe_allow_html=True)
    c7, c8, c9, c10 = st.columns(4)
    los_ctrl    = c7.number_input("Control — Mean / Median (days)", value=float(d.get("los_control") or 0),   step=0.1, format="%.2f", key=f"{outcome}_{study_idx}_lc")
    los_ctrl_sd = c8.number_input("Control — SD",                   value=float(d.get("los_ctrl_sd") or 0),   step=0.1, format="%.2f", key=f"{outcome}_{study_idx}_lcs")
    ctrl_iqr_lo = c9.number_input("Control — IQR Lower",            value=float(d.get("ctrl_iqr_lower") or 0), step=0.1, format="%.2f", key=f"{outcome}_{study_idx}_lcil")
    ctrl_iqr_hi = c10.number_input("Control — IQR Upper",           value=float(d.get("ctrl_iqr_upper") or 0), step=0.1, format="%.2f", key=f"{outcome}_{study_idx}_lciu")

    # Derived
    c11, c12, c13, c14 = st.columns(4)
    mean_diff   = c11.number_input("Mean/Median Difference (days)", value=float(d.get("mean_diff") or 0), step=0.01, format="%.3f", key=f"{outcome}_{study_idx}_md")
    pct_red     = c12.number_input("% Reduction in LOS",            value=float(d.get("pct_reduction") or 0), step=0.1, format="%.1f", key=f"{outcome}_{study_idx}_pr")
    outliers    = c13.selectbox("Outliers Addressed",  YN,
                                 index=YN.index(d.get("outliers_addressed","")) if d.get("outliers_addressed","") in YN else 0,
                                 key=f"{outcome}_{study_idx}_out")
    los_def     = c14.text_input("LOS Definition", d.get("los_definition",""),
                                  key=f"{outcome}_{study_idx}_ldef",
                                  help="'days to discharge' or 'fit for discharge'?")

    meta = meta_fields(f"{outcome}_{study_idx}", d)

    col_sv, col_cl = st.columns([1, 5])
    if col_sv.button("💾 Save Record", key=f"save_{outcome}_{study_idx}", type="primary"):
        rec = {**id_data,
               "los_type": los_type, "los_measure": los_measure,
               "los_eras": los_eras, "los_eras_sd": los_eras_sd,
               "eras_iqr_lower": eras_iqr_lo, "eras_iqr_upper": eras_iqr_hi,
               "los_control": los_ctrl, "los_ctrl_sd": los_ctrl_sd,
               "ctrl_iqr_lower": ctrl_iqr_lo, "ctrl_iqr_upper": ctrl_iqr_hi,
               "mean_diff": mean_diff, "pct_reduction": pct_red,
               "outliers_addressed": outliers, "los_definition": los_def,
               **meta}
        save_record(study_idx, outcome, rec, ss.edit_record.get(outcome))
        ss.edit_record[outcome] = None
        ss.edit_state[outcome] = {}
        ss.excel_bytes = None
        st.success("Saved!"); st.rerun()
    if not is_new and col_cl.button("Cancel Edit", key=f"cancel_{outcome}_{study_idx}"):
        ss.edit_record[outcome] = None; ss.edit_state[outcome] = {}; st.rerun()


# ── O2_COMP form ─────────────────────────────────────────────────────────────
def form_o2(study_idx: int):
    outcome = "O2_COMP"
    init_edit(outcome)
    records_panel(study_idx, outcome)

    is_new = ss.edit_record.get(outcome) is None
    st.markdown(f"### {'➕ New Complication Record' if is_new else f'✏️ Editing Record #{ss.edit_record[outcome]+1}'}")
    d = ss.edit_state.get(outcome, {})
    id_data = common_id_form(f"{outcome}_{study_idx}", d)

    st.markdown("**Complication Data**")
    c1, c2, c3, c4 = st.columns(4)
    comp_scope  = c1.selectbox("Complication Scope", COMP_SCOPES,
                                index=COMP_SCOPES.index(d.get("complication_scope","")) if d.get("complication_scope","") in COMP_SCOPES else 0,
                                key=f"{outcome}_{study_idx}_cs")
    comp_type   = c2.text_input("Complication Type / Name", d.get("complication_type",""),
                                 key=f"{outcome}_{study_idx}_ct",
                                 help="e.g. 'SSI', 'Anastomotic leak', 'Overall'")
    cd_reported = c3.selectbox("Clavien-Dindo Reported", YN,
                                index=YN.index(d.get("cd_reported","")) if d.get("cd_reported","") in YN else 0,
                                key=f"{outcome}_{study_idx}_cdr")
    cd_grade    = c4.selectbox("Clavien-Dindo Grade", CLAVIEN_GRADES,
                                index=CLAVIEN_GRADES.index(d.get("cd_grade","")) if d.get("cd_grade","") in CLAVIEN_GRADES else 0,
                                key=f"{outcome}_{study_idx}_cdg")
    c5,c6,c7,c8 = st.columns(4)
    ev_eras     = c5.number_input("Events ERAS (n)",      value=int(d.get("events_eras") or 0),    min_value=0, key=f"{outcome}_{study_idx}_ee")
    ev_ctrl     = c6.number_input("Events Control (n)",   value=int(d.get("events_control") or 0), min_value=0, key=f"{outcome}_{study_idx}_ec")
    rate_eras   = c7.number_input("Rate ERAS (%)",        value=float(d.get("rate_eras") or 0),    step=0.1, format="%.2f", key=f"{outcome}_{study_idx}_re")
    rate_ctrl   = c8.number_input("Rate Control (%)",     value=float(d.get("rate_ctrl") or 0),    step=0.1, format="%.2f", key=f"{outcome}_{study_idx}_rc")

    st.markdown("**Specific Complication Types Reported**")
    c9,c10,c11,c12,c13 = st.columns(5)
    ssi_rep  = c9.selectbox("SSI",            YN, index=YN.index(d.get("ssi_reported","")) if d.get("ssi_reported","") in YN else 0, key=f"{outcome}_{study_idx}_ssi")
    al_rep   = c10.selectbox("Anastomotic Leak", YN, index=YN.index(d.get("al_reported","")) if d.get("al_reported","") in YN else 0, key=f"{outcome}_{study_idx}_al")
    poi_rep  = c11.selectbox("Ileus / POI",   YN, index=YN.index(d.get("poi_reported","")) if d.get("poi_reported","") in YN else 0, key=f"{outcome}_{study_idx}_poi")
    pulm_rep = c12.selectbox("Pulmonary",      YN, index=YN.index(d.get("pulm_reported","")) if d.get("pulm_reported","") in YN else 0, key=f"{outcome}_{study_idx}_pulm")
    urin_rep = c13.selectbox("Urinary",        YN, index=YN.index(d.get("urin_reported","")) if d.get("urin_reported","") in YN else 0, key=f"{outcome}_{study_idx}_urin")

    meta = meta_fields(f"{outcome}_{study_idx}", d)

    if st.button("💾 Save Record", key=f"save_{outcome}_{study_idx}", type="primary"):
        rec = {**id_data,
               "complication_scope": comp_scope, "complication_type": comp_type,
               "cd_reported": cd_reported, "cd_grade": cd_grade,
               "events_eras": int(ev_eras), "events_control": int(ev_ctrl),
               "rate_eras": rate_eras, "rate_ctrl": rate_ctrl,
               "ssi_reported": ssi_rep, "al_reported": al_rep,
               "poi_reported": poi_rep, "pulm_reported": pulm_rep,
               "urin_reported": urin_rep, **meta}
        save_record(study_idx, outcome, rec, ss.edit_record.get(outcome))
        ss.edit_record[outcome] = None; ss.edit_state[outcome] = {}
        ss.excel_bytes = None; st.success("Saved!"); st.rerun()


# ── O3_READ form ──────────────────────────────────────────────────────────────
def form_o3(study_idx: int):
    outcome = "O3_READ"
    init_edit(outcome)
    records_panel(study_idx, outcome)
    d = ss.edit_state.get(outcome, {})
    id_data = common_id_form(f"{outcome}_{study_idx}", d)

    st.markdown("**Readmission Data**")
    c1,c2,c3,c4 = st.columns(4)
    timeframe   = c1.selectbox("Timeframe", READ_TIMEFRAMES,
                                index=READ_TIMEFRAMES.index(d.get("readmission_timeframe","")) if d.get("readmission_timeframe","") in READ_TIMEFRAMES else 0,
                                key=f"{outcome}_{study_idx}_tf")
    tf_other    = c2.text_input("Other Timeframe (days)", d.get("timeframe_other",""), key=f"{outcome}_{study_idx}_tfo")
    ev_eras     = c3.number_input("Events ERAS (n)",     value=int(d.get("events_eras") or 0), min_value=0, key=f"{outcome}_{study_idx}_ee")
    ev_ctrl     = c4.number_input("Events Control (n)",  value=int(d.get("events_control") or 0), min_value=0, key=f"{outcome}_{study_idx}_ec")
    c5,c6,c7,c8 = st.columns(4)
    rate_eras   = c5.number_input("Rate ERAS (%)",       value=float(d.get("readmit_rate_eras") or 0), step=0.1, format="%.2f", key=f"{outcome}_{study_idx}_re")
    rate_ctrl   = c6.number_input("Rate Control (%)",    value=float(d.get("readmit_rate_ctrl") or 0), step=0.1, format="%.2f", key=f"{outcome}_{study_idx}_rc")
    cause_rep   = c7.selectbox("Cause Reported", YN,
                                index=YN.index(d.get("cause_reported","")) if d.get("cause_reported","") in YN else 0,
                                key=f"{outcome}_{study_idx}_cr")
    unplanned   = c8.selectbox("Unplanned Only", YN,
                                index=YN.index(d.get("unplanned_only","")) if d.get("unplanned_only","") in YN else 0,
                                key=f"{outcome}_{study_idx}_up")
    top_cause   = st.text_input("Top Readmission Cause", d.get("top_cause",""), key=f"{outcome}_{study_idx}_tc")
    meta = meta_fields(f"{outcome}_{study_idx}", d)

    if st.button("💾 Save Record", key=f"save_{outcome}_{study_idx}", type="primary"):
        rec = {**id_data,
               "readmission_timeframe": timeframe, "timeframe_other": tf_other,
               "events_eras": int(ev_eras), "events_control": int(ev_ctrl),
               "readmit_rate_eras": rate_eras, "readmit_rate_ctrl": rate_ctrl,
               "cause_reported": cause_rep, "top_cause": top_cause,
               "unplanned_only": unplanned, **meta}
        save_record(study_idx, outcome, rec, None)
        ss.excel_bytes = None; st.success("Saved!"); st.rerun()


# ── O4_MORT form ──────────────────────────────────────────────────────────────
def form_o4(study_idx: int):
    outcome = "O4_MORT"
    init_edit(outcome)
    records_panel(study_idx, outcome)
    d = ss.edit_state.get(outcome, {})
    id_data = common_id_form(f"{outcome}_{study_idx}", d)

    st.markdown("**Mortality Data**")
    c1,c2,c3,c4 = st.columns(4)
    mort_type   = c1.selectbox("Mortality Type", MORT_TYPES,
                                index=MORT_TYPES.index(d.get("mortality_type","")) if d.get("mortality_type","") in MORT_TYPES else 0,
                                key=f"{outcome}_{study_idx}_mt")
    ev_eras     = c2.number_input("Events ERAS (n)",     value=int(d.get("events_eras") or 0), min_value=0, key=f"{outcome}_{study_idx}_ee",
                                   help="Enter 0 (not blank) for zero events")
    ev_ctrl     = c3.number_input("Events Control (n)",  value=int(d.get("events_control") or 0), min_value=0, key=f"{outcome}_{study_idx}_ec")
    op_vs_postop= c4.text_input("Operative vs Non-operative",  d.get("operative_vs_nonop",""), key=f"{outcome}_{study_idx}_ovn")
    c5,c6,c7 = st.columns(3)
    rate_eras   = c5.number_input("Mortality Rate ERAS (%)",    value=float(d.get("mort_rate_eras") or 0), step=0.01, format="%.3f", key=f"{outcome}_{study_idx}_re")
    rate_ctrl   = c6.number_input("Mortality Rate Control (%)", value=float(d.get("mort_rate_ctrl") or 0), step=0.01, format="%.3f", key=f"{outcome}_{study_idx}_rc")
    cause_rep   = c7.selectbox("Cause of Death Reported", YN,
                                index=YN.index(d.get("cause_reported","")) if d.get("cause_reported","") in YN else 0,
                                key=f"{outcome}_{study_idx}_cr")
    meta = meta_fields(f"{outcome}_{study_idx}", d)

    if st.button("💾 Save Record", key=f"save_{outcome}_{study_idx}", type="primary"):
        rec = {**id_data,
               "mortality_type": mort_type,
               "events_eras": int(ev_eras), "events_control": int(ev_ctrl),
               "mort_rate_eras": rate_eras, "mort_rate_ctrl": rate_ctrl,
               "cause_reported": cause_rep, "operative_vs_nonop": op_vs_postop,
               **meta}
        save_record(study_idx, outcome, rec, None)
        ss.excel_bytes = None; st.success("Saved!"); st.rerun()


# ── O5_COST form ──────────────────────────────────────────────────────────────
def form_o5(study_idx: int):
    outcome = "O5_COST"
    init_edit(outcome)
    records_panel(study_idx, outcome)
    d = ss.edit_state.get(outcome, {})
    id_data = common_id_form(f"{outcome}_{study_idx}", d)

    st.markdown("**Cost-Effectiveness Data**")
    c1,c2,c3,c4 = st.columns(4)
    persp       = c1.selectbox("Cost Perspective", COST_PERSP,
                                index=COST_PERSP.index(d.get("cost_perspective","")) if d.get("cost_perspective","") in COST_PERSP else 0,
                                key=f"{outcome}_{study_idx}_cp")
    components  = c2.selectbox("Cost Components",  COST_COMP,
                                index=COST_COMP.index(d.get("cost_components","")) if d.get("cost_components","") in COST_COMP else 0,
                                key=f"{outcome}_{study_idx}_cc")
    cost_year   = c3.number_input("Year of Costing", min_value=1990, max_value=2026,
                                   value=int(d.get("year_of_costing") or 2020), key=f"{outcome}_{study_idx}_cy")
    currency    = c4.text_input("Original Currency", d.get("original_currency",""),
                                 key=f"{outcome}_{study_idx}_cur", help="e.g. USD, EUR, ETB, GBP")
    c5,c6,c7 = st.columns(3)
    cost_eras   = c5.number_input("Total Cost ERAS (orig.)",    value=float(d.get("cost_eras_orig") or 0),   step=0.01, format="%.2f", key=f"{outcome}_{study_idx}_ce")
    cost_ctrl   = c6.number_input("Total Cost Control (orig.)", value=float(d.get("cost_ctrl_orig") or 0),   step=0.01, format="%.2f", key=f"{outcome}_{study_idx}_cc2")
    cost_diff   = c7.number_input("Cost Difference (orig.)",    value=float(d.get("cost_diff_orig") or 0),   step=0.01, format="%.2f", key=f"{outcome}_{study_idx}_cd")
    st.markdown("**PPP-Adjusted Costs (USD 2023)**")
    c8,c9,c10,c11,c12 = st.columns(5)
    ppp_rate    = c8.number_input("PPP Conversion Rate",       value=float(d.get("ppp_rate") or 0),         step=0.001, format="%.4f", key=f"{outcome}_{study_idx}_ppp")
    cost_eras_u = c9.number_input("ERAS USD 2023",             value=float(d.get("cost_eras_usd") or 0),    step=0.01, format="%.2f",  key=f"{outcome}_{study_idx}_ceu")
    cost_ctrl_u = c10.number_input("Control USD 2023",         value=float(d.get("cost_ctrl_usd") or 0),    step=0.01, format="%.2f",  key=f"{outcome}_{study_idx}_ccu")
    cost_diff_u = c11.number_input("Difference USD 2023",      value=float(d.get("cost_diff_usd") or 0),    step=0.01, format="%.2f",  key=f"{outcome}_{study_idx}_cdu")
    pct_save    = c12.number_input("% Cost Reduction",         value=float(d.get("pct_cost_reduction") or 0), step=0.1, format="%.1f", key=f"{outcome}_{study_idx}_pcs")
    c13,c14,c15,c16 = st.columns(4)
    icer_rep    = c13.selectbox("ICER Reported", YN,
                                 index=YN.index(d.get("icer_reported","")) if d.get("icer_reported","") in YN else 0,
                                 key=f"{outcome}_{study_idx}_icr")
    icer_val    = c14.text_input("ICER Value",   d.get("icer_value",""), key=f"{outcome}_{study_idx}_icv")
    bedday_rep  = c15.selectbox("Cost per Bed-Day", YN,
                                 index=YN.index(d.get("cost_per_bedday","")) if d.get("cost_per_bedday","") in YN else 0,
                                 key=f"{outcome}_{study_idx}_cbd")
    sens_rep    = c16.selectbox("Sensitivity Analysis", YN,
                                 index=YN.index(d.get("sensitivity_analysis","")) if d.get("sensitivity_analysis","") in YN else 0,
                                 key=f"{outcome}_{study_idx}_sa")
    meta = meta_fields(f"{outcome}_{study_idx}", d)

    if st.button("💾 Save Record", key=f"save_{outcome}_{study_idx}", type="primary"):
        rec = {**id_data,
               "cost_perspective": persp, "cost_components": components,
               "year_of_costing": int(cost_year), "original_currency": currency,
               "cost_eras_orig": cost_eras, "cost_ctrl_orig": cost_ctrl,
               "cost_diff_orig": cost_diff, "ppp_rate": ppp_rate,
               "cost_eras_usd": cost_eras_u, "cost_ctrl_usd": cost_ctrl_u,
               "cost_diff_usd": cost_diff_u, "pct_cost_reduction": pct_save,
               "icer_reported": icer_rep, "icer_value": icer_val,
               "cost_per_bedday": bedday_rep, "sensitivity_analysis": sens_rep,
               **meta}
        save_record(study_idx, outcome, rec, None)
        ss.excel_bytes = None; st.success("Saved!"); st.rerun()


# ── ADHERENCE form ────────────────────────────────────────────────────────────
def form_adherence(study_idx: int):
    outcome = "ADHERENCE"
    init_edit(outcome)
    records_panel(study_idx, outcome)
    d = ss.edit_state.get(outcome, {})
    id_data = common_id_form(f"{outcome}_{study_idx}", d)

    st.markdown("**Overall Adherence**")
    c1,c2,c3 = st.columns(3)
    total_elem  = c1.number_input("Total Protocol Elements (n)",    value=int(d.get("total_elements") or 0),     min_value=0, key=f"{outcome}_{study_idx}_te")
    impl_elem   = c2.number_input("Total Implemented Elements (n)", value=int(d.get("implemented_elements") or 0), min_value=0, key=f"{outcome}_{study_idx}_ie")
    overall_adh = c3.number_input("Overall Adherence Rate (%)",     value=float(d.get("overall_adherence_rate") or 0), step=0.1, format="%.1f", key=f"{outcome}_{study_idx}_oa")
    c4,c5 = st.columns(2)
    adh_cat     = c4.selectbox("Adherence Category", ADHERENCE_CATS,
                                index=ADHERENCE_CATS.index(d.get("adherence_category","")) if d.get("adherence_category","") in ADHERENCE_CATS else 0,
                                key=f"{outcome}_{study_idx}_ac")
    adh_method  = c5.selectbox("Measurement Method", ADHERENCE_METHODS,
                                index=ADHERENCE_METHODS.index(d.get("adherence_method","")) if d.get("adherence_method","") in ADHERENCE_METHODS else 0,
                                key=f"{outcome}_{study_idx}_am")
    c6,c7,c8 = st.columns(3)
    preop_adh   = c6.number_input("Pre-op Phase Adherence (%)",   value=float(d.get("preop_phase_adh") or 0),  step=0.1, format="%.1f", key=f"{outcome}_{study_idx}_poa")
    intraop_adh = c7.number_input("Intra-op Phase Adherence (%)", value=float(d.get("intraop_phase_adh") or 0), step=0.1, format="%.1f", key=f"{outcome}_{study_idx}_ioa")
    postop_adh  = c8.number_input("Post-op Phase Adherence (%)",  value=float(d.get("postop_phase_adh") or 0), step=0.1, format="%.1f", key=f"{outcome}_{study_idx}_popa")

    st.markdown("**Element-Level Adherence** — Enter: 0 = Not implemented · 1 = Fully implemented · 0.5 = Partial · or % if reported")
    elem_vals = {}
    cols = st.columns(3)
    for i, (key, label) in enumerate(ERAS_ELEMENTS):
        val = cols[i % 3].number_input(label, value=float(d.get(key) or 0),
                                        step=0.5, min_value=0.0, max_value=100.0,
                                        format="%.1f", key=f"{outcome}_{study_idx}_{key}")
        elem_vals[key] = val

    st.markdown("**Adherence-Outcome Association**")
    c9,c10,c11 = st.columns(3)
    los_assoc   = c9.number_input("LOS Difference Associated (days)", value=float(d.get("los_diff_assoc") or 0), step=0.1, format="%.2f", key=f"{outcome}_{study_idx}_lda")
    comp_assoc  = c10.number_input("Complication OR/RR",              value=float(d.get("comp_or_rr") or 0),     step=0.001, format="%.4f", key=f"{outcome}_{study_idx}_cor")
    p_assoc     = c11.number_input("p-value (Adherence-Outcome)",     value=float(d.get("p_assoc") or 0),        step=0.001, format="%.4f", key=f"{outcome}_{study_idx}_pao")
    c12,c13,c14 = st.columns(3)
    thresh_rep  = c12.selectbox("Threshold Reported", YN,
                                 index=YN.index(d.get("threshold_reported","")) if d.get("threshold_reported","") in YN else 0,
                                 key=f"{outcome}_{study_idx}_thr")
    thresh_val  = c13.number_input("Threshold Value (%)", value=float(d.get("threshold_value") or 0), step=0.1, format="%.1f", key=f"{outcome}_{study_idx}_thv")
    dose_resp   = c14.selectbox("Dose-Response Analysis", YN,
                                 index=YN.index(d.get("dose_response","")) if d.get("dose_response","") in YN else 0,
                                 key=f"{outcome}_{study_idx}_dr")
    notes = st.text_area("Notes", d.get("notes",""), height=80, key=f"{outcome}_{study_idx}_notes")

    if st.button("💾 Save Record", key=f"save_{outcome}_{study_idx}", type="primary"):
        rec = {**id_data,
               "total_elements": int(total_elem), "implemented_elements": int(impl_elem),
               "overall_adherence_rate": overall_adh, "adherence_category": adh_cat,
               "adherence_method": adh_method, "preop_phase_adh": preop_adh,
               "intraop_phase_adh": intraop_adh, "postop_phase_adh": postop_adh,
               **elem_vals,
               "los_diff_assoc": los_assoc, "comp_or_rr": comp_assoc,
               "p_assoc": p_assoc, "threshold_reported": thresh_rep,
               "threshold_value": thresh_val, "dose_response": dose_resp,
               "notes": notes}
        save_record(study_idx, outcome, rec, None)
        ss.excel_bytes = None; st.success("Saved!"); st.rerun()


# ── GEO_DISPARITY form ────────────────────────────────────────────────────────
def form_geo(study_idx: int):
    outcome = "GEO_DISPARITY"
    init_edit(outcome)
    records_panel(study_idx, outcome)
    d = ss.edit_state.get(outcome, {})
    # Geo uses its own ID block (no Specific Procedure / Study Design in the sheet)
    with st.expander("📌 Study Identification", expanded=True):
        c1,c2,c3,c4 = st.columns(4)
        study_id = c1.text_input("Study ID", d.get("study_id",""), key=f"{outcome}_{study_idx}_sid")
        auth     = c2.text_input("First Author", d.get("first_author",""), key=f"{outcome}_{study_idx}_auth")
        year     = c3.number_input("Year", 1990, 2030, int(d.get("year",2020)), key=f"{outcome}_{study_idx}_yr")
        country  = c4.text_input("Country", d.get("country",""), key=f"{outcome}_{study_idx}_cnt")
        c5,c6,c7,c8 = st.columns(4)
        who_reg  = c5.selectbox("WHO Region", WHO_REGIONS,
                                 index=WHO_REGIONS.index(d.get("who_region","")) if d.get("who_region","") in WHO_REGIONS else 0,
                                 key=f"{outcome}_{study_idx}_who")
        inc_grp  = c6.selectbox("Income Group", INCOME_GROUPS,
                                 index=INCOME_GROUPS.index(d.get("income_group","")) if d.get("income_group","") in INCOME_GROUPS else 0,
                                 key=f"{outcome}_{study_idx}_inc")
        subreg   = c7.selectbox("Sub-region", SUBREGS,
                                 index=SUBREGS.index(d.get("subregion","")) if d.get("subregion","") in SUBREGS else 0,
                                 key=f"{outcome}_{study_idx}_sr")
        ur_class = c8.selectbox("Urban/Rural", URBAN_RURAL,
                                 index=URBAN_RURAL.index(d.get("urban_rural","")) if d.get("urban_rural","") in URBAN_RURAL else 0,
                                 key=f"{outcome}_{study_idx}_ur")
        c9,c10,c11 = st.columns(3)
        hosp_lev = c9.selectbox("Hospital Level", HOSPITAL_LEVELS,
                                 index=HOSPITAL_LEVELS.index(d.get("hospital_level","")) if d.get("hospital_level","") in HOSPITAL_LEVELS else 0,
                                 key=f"{outcome}_{study_idx}_hl")
        specialty= c10.selectbox("Surgical Specialty", SPECIALTIES,
                                  index=SPECIALTIES.index(d.get("surgical_specialty","")) if d.get("surgical_specialty","") in SPECIALTIES else 0,
                                  key=f"{outcome}_{study_idx}_sp")
    id_data = dict(study_id=study_id, first_author=auth, year=int(year),
                   country=country, who_region=who_reg, income_group=inc_grp,
                   subregion=subreg, urban_rural=ur_class,
                   hospital_level=hosp_lev, surgical_specialty=specialty)

    st.markdown("**Country-Level Contextual Variables** *(from World Bank for study year)*")
    c1,c2,c3,c4,c5 = st.columns(5)
    gdp       = c1.number_input("GDP per Capita (USD)",          value=float(d.get("gdp_per_capita") or 0),      step=1.0, format="%.0f", key=f"{outcome}_{study_idx}_gdp")
    hlth_exp  = c2.number_input("Health Expenditure per Capita", value=float(d.get("health_exp_per_cap") or 0),  step=1.0, format="%.0f", key=f"{outcome}_{study_idx}_he")
    surg_dens = c3.number_input("Surgical Workforce /100k",      value=float(d.get("surg_workforce_per100k") or 0), step=0.1, format="%.2f", key=f"{outcome}_{study_idx}_sw")
    ana_dens  = c4.number_input("Anaesthetist Density /100k",    value=float(d.get("anaesthetist_per100k") or 0),   step=0.1, format="%.2f", key=f"{outcome}_{study_idx}_ad")
    icu_beds  = c5.number_input("ICU Beds /100k",                value=float(d.get("icu_beds_per100k") or 0),        step=0.1, format="%.2f", key=f"{outcome}_{study_idx}_ib")

    st.markdown("**ERAS Outcomes in this Study**")
    c6,c7,c8,c9,c10 = st.columns(5)
    adh_rate   = c6.number_input("ERAS Adherence Rate (%)", value=float(d.get("overall_adherence_rate") or 0), step=0.1, format="%.1f", key=f"{outcome}_{study_idx}_ar")
    los_eras   = c7.number_input("LOS ERAS (days)",         value=float(d.get("los_eras") or 0), step=0.1, format="%.2f", key=f"{outcome}_{study_idx}_le")
    los_ctrl   = c8.number_input("LOS Control (days)",      value=float(d.get("los_ctrl") or 0), step=0.1, format="%.2f", key=f"{outcome}_{study_idx}_lc")
    los_diff   = c9.number_input("LOS Difference (days)",   value=float(d.get("los_diff") or 0), step=0.1, format="%.2f", key=f"{outcome}_{study_idx}_ld")
    comp_eras  = c10.number_input("Complication Rate ERAS (%)",  value=float(d.get("comp_rate_eras") or 0), step=0.1, format="%.2f", key=f"{outcome}_{study_idx}_cre")
    c11,c12,c13,c14,c15 = st.columns(5)
    comp_ctrl  = c11.number_input("Complication Rate Ctrl (%)",  value=float(d.get("comp_rate_ctrl") or 0), step=0.1, format="%.2f", key=f"{outcome}_{study_idx}_crc")
    comp_rror  = c12.number_input("RR/OR Complications",         value=float(d.get("rr_or_comp") or 0),     step=0.001, format="%.4f", key=f"{outcome}_{study_idx}_rro")
    read_eras  = c13.number_input("Readmission Rate ERAS (%)",   value=float(d.get("readmit_rate_eras") or 0), step=0.1, format="%.2f", key=f"{outcome}_{study_idx}_rre")
    mort_eras  = c14.number_input("Mortality Rate ERAS (%)",     value=float(d.get("mort_rate_eras") or 0),   step=0.01, format="%.3f", key=f"{outcome}_{study_idx}_mre")
    cost_sav_r = c15.selectbox("Cost Savings Reported", YN,
                                index=YN.index(d.get("cost_savings_reported","")) if d.get("cost_savings_reported","") in YN else 0,
                                key=f"{outcome}_{study_idx}_csr")
    cost_sav   = st.number_input("Cost Savings (USD PPP)", value=float(d.get("cost_savings_usd") or 0), step=0.01, format="%.2f", key=f"{outcome}_{study_idx}_cs")

    st.markdown("**Barriers & Enablers**")
    bar_rep = st.selectbox("Barriers Reported", YN,
                            index=YN.index(d.get("barriers_reported","")) if d.get("barriers_reported","") in YN else 0,
                            key=f"{outcome}_{study_idx}_br")
    bc1,bc2,bc3,bc4,bc5 = st.columns(5)
    b_infra  = bc1.selectbox("Infrastructure",  YN, index=YN.index(d.get("b_infrastructure","")) if d.get("b_infrastructure","") in YN else 0, key=f"{outcome}_{study_idx}_bi")
    b_train  = bc2.selectbox("Staff Training",  YN, index=YN.index(d.get("b_training","")) if d.get("b_training","") in YN else 0, key=f"{outcome}_{study_idx}_bt")
    b_supply = bc3.selectbox("Supply Chain",    YN, index=YN.index(d.get("b_supply","")) if d.get("b_supply","") in YN else 0, key=f"{outcome}_{study_idx}_bs")
    b_cult   = bc4.selectbox("Cultural/Patient",YN, index=YN.index(d.get("b_cultural","")) if d.get("b_cultural","") in YN else 0, key=f"{outcome}_{study_idx}_bc")
    b_fin    = bc5.selectbox("Financial",       YN, index=YN.index(d.get("b_financial","")) if d.get("b_financial","") in YN else 0, key=f"{outcome}_{study_idx}_bf")
    b_other  = st.text_input("Other Barrier (specify)", d.get("b_other",""), key=f"{outcome}_{study_idx}_bo")

    en_rep = st.selectbox("Enablers Reported", YN,
                           index=YN.index(d.get("enablers_reported","")) if d.get("enablers_reported","") in YN else 0,
                           key=f"{outcome}_{study_idx}_er")
    ec1,ec2,ec3 = st.columns(3)
    e_lead   = ec1.selectbox("Leadership",        YN, index=YN.index(d.get("e_leadership","")) if d.get("e_leadership","") in YN else 0, key=f"{outcome}_{study_idx}_el")
    e_train  = ec2.selectbox("Training Programme",YN, index=YN.index(d.get("e_training","")) if d.get("e_training","") in YN else 0, key=f"{outcome}_{study_idx}_et")
    e_proto  = ec3.selectbox("Protocol Adaptation",YN, index=YN.index(d.get("e_protocol","")) if d.get("e_protocol","") in YN else 0, key=f"{outcome}_{study_idx}_ep")

    c_dose,c_coef = st.columns(2)
    dose_an = c_dose.selectbox("Dose-Response Analysis", YN,
                                index=YN.index(d.get("dose_resp_analysis","")) if d.get("dose_resp_analysis","") in YN else 0,
                                key=f"{outcome}_{study_idx}_dra")
    dose_co = c_coef.number_input("Dose-Response Coefficient", value=float(d.get("dose_resp_coef") or 0), step=0.0001, format="%.6f", key=f"{outcome}_{study_idx}_drc")
    notes = st.text_area("Notes", d.get("notes",""), height=80, key=f"{outcome}_{study_idx}_notes")

    if st.button("💾 Save Record", key=f"save_{outcome}_{study_idx}", type="primary"):
        rec = {**id_data,
               "gdp_per_capita": gdp, "health_exp_per_cap": hlth_exp,
               "surg_workforce_per100k": surg_dens, "anaesthetist_per100k": ana_dens,
               "icu_beds_per100k": icu_beds, "overall_adherence_rate": adh_rate,
               "los_eras": los_eras, "los_ctrl": los_ctrl, "los_diff": los_diff,
               "comp_rate_eras": comp_eras, "comp_rate_ctrl": comp_ctrl,
               "rr_or_comp": comp_rror, "readmit_rate_eras": read_eras,
               "mort_rate_eras": mort_eras, "cost_savings_reported": cost_sav_r,
               "cost_savings_usd": cost_sav, "barriers_reported": bar_rep,
               "b_infrastructure": b_infra, "b_training": b_train,
               "b_supply": b_supply, "b_cultural": b_cult, "b_financial": b_fin,
               "b_other": b_other, "enablers_reported": en_rep,
               "e_leadership": e_lead, "e_training": e_train, "e_protocol": e_proto,
               "dose_resp_analysis": dose_an, "dose_resp_coef": dose_co,
               "notes": notes}
        save_record(study_idx, outcome, rec, None)
        ss.excel_bytes = None; st.success("Saved!"); st.rerun()


# ── Excel export ──────────────────────────────────────────────────────────────
HEADER_ROW = PatternFill("solid", fgColor="1B3A6B")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SUBHEAD_FILL = PatternFill("solid", fgColor="2E6DA4")
SUBHEAD_FONT = Font(bold=True, color="FFFFFF", size=9)
ALT_FILL     = PatternFill("solid", fgColor="EBF3FB")
BODY_FONT    = Font(size=9)
CENT_ALIGN   = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _write_sheet(ws, headers: list, records: list, sheet_title: str):
    ws.freeze_panes = "A4"
    # Row 1: sheet title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    cell = ws.cell(1, 1, sheet_title)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor="0D1B2A")
    cell.alignment = CENT_ALIGN
    ws.row_dimensions[1].height = 22
    # Row 3: column headers
    for ci, h in enumerate(headers, 1):
        c = ws.cell(3, ci, h)
        c.font = HEADER_FONT; c.fill = HEADER_ROW; c.alignment = CENT_ALIGN
    ws.row_dimensions[3].height = 40
    # Data rows
    for ri, rec in enumerate(records, 4):
        fill = ALT_FILL if ri % 2 == 0 else None
        for ci, h in enumerate(headers, 1):
            val = rec.get(h, rec.get(h.lower().replace(" ","_").replace("/","_").replace("(","").replace(")",""), ""))
            c = ws.cell(ri, ci, val)
            c.font = BODY_FONT; c.alignment = LEFT_ALIGN
            if fill: c.fill = fill
    # Column widths
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = max(12, min(28, len(h)+2))


# Column header lists matching the workbook exactly
O1_COLS = [
    "Study ID","First Author","Year","Country","WHO Region","Income Group",
    "Surgical Specialty","Specific Procedure","Study Design","N Total","N ERAS","N Control",
    "Adherence Rate (%)","Implementation Type",
    "LOS Type (Total Hospital/ICU/Ward)","LOS Measure (Mean/Median)",
    "LOS — ERAS (days)","LOS ERAS — SD","LOS ERAS — IQR Lower","LOS ERAS — IQR Upper",
    "LOS — Control (days)","LOS Control — SD","LOS Control — IQR Lower","LOS Control — IQR Upper",
    "Mean/Median Difference (days)","% Reduction in LOS","Outliers Addressed (Y/N)",
    "LOS Definition (days to discharge/fit)",
    "Effect Measure (MD/RR/OR/HR/RD)","Effect Size","95% CI Lower","95% CI Upper",
    "p-value","Heterogeneity I2 (%)","Tau2","Statistical Method",
    "Imputation Used (Y/N)","Notes","Extracted By","Date Extracted","Verified By"
]
O1_MAP = {  # header → record key
    "LOS Type (Total Hospital/ICU/Ward)":      "los_type",
    "LOS Measure (Mean/Median)":               "los_measure",
    "LOS — ERAS (days)":                       "los_eras",
    "LOS ERAS — SD":                           "los_eras_sd",
    "LOS ERAS — IQR Lower":                    "eras_iqr_lower",
    "LOS ERAS — IQR Upper":                    "eras_iqr_upper",
    "LOS — Control (days)":                    "los_control",
    "LOS Control — SD":                        "los_ctrl_sd",
    "LOS Control — IQR Lower":                 "ctrl_iqr_lower",
    "LOS Control — IQR Upper":                 "ctrl_iqr_upper",
    "Mean/Median Difference (days)":           "mean_diff",
    "% Reduction in LOS":                      "pct_reduction",
    "Outliers Addressed (Y/N)":                "outliers_addressed",
    "LOS Definition (days to discharge/fit)":  "los_definition",
}

O2_COLS = [
    "Study ID","First Author","Year","Country","WHO Region","Income Group",
    "Surgical Specialty","Specific Procedure","Study Design","N Total","N ERAS","N Control",
    "Adherence Rate (%)","Implementation Type",
    "Complication Scope (Overall/Specific)","Complication Type / Name",
    "Clavien-Dindo Grade Reported (Y/N)","Clavien-Dindo Grade (I/II/III/IV/V)",
    "Events ERAS (n)","Events Control (n)",
    "Complication Rate ERAS (%)","Complication Rate Control (%)",
    "SSI Reported (Y/N)","Anastomotic Leak Reported (Y/N)",
    "Ileus / POI Reported (Y/N)","Pulmonary Complications Reported (Y/N)",
    "Urinary Complications Reported (Y/N)",
    "Effect Measure (MD/RR/OR/HR/RD)","Effect Size","95% CI Lower","95% CI Upper",
    "p-value","Heterogeneity I2 (%)","Tau2","Statistical Method",
    "Imputation Used (Y/N)","Notes","Extracted By","Date Extracted","Verified By"
]
O2_MAP = {
    "Complication Scope (Overall/Specific)":   "complication_scope",
    "Complication Type / Name":                "complication_type",
    "Clavien-Dindo Grade Reported (Y/N)":      "cd_reported",
    "Clavien-Dindo Grade (I/II/III/IV/V)":     "cd_grade",
    "Events ERAS (n)":                         "events_eras",
    "Events Control (n)":                      "events_control",
    "Complication Rate ERAS (%)":              "rate_eras",
    "Complication Rate Control (%)":           "rate_ctrl",
    "SSI Reported (Y/N)":                      "ssi_reported",
    "Anastomotic Leak Reported (Y/N)":         "al_reported",
    "Ileus / POI Reported (Y/N)":              "poi_reported",
    "Pulmonary Complications Reported (Y/N)":  "pulm_reported",
    "Urinary Complications Reported (Y/N)":    "urin_reported",
}

O3_COLS = [
    "Study ID","First Author","Year","Country","WHO Region","Income Group",
    "Surgical Specialty","Specific Procedure","Study Design","N Total","N ERAS","N Control",
    "Adherence Rate (%)","Implementation Type",
    "Readmission Timeframe (30-day/90-day/1-year/Other)","Timeframe — Other (specify days)",
    "Events ERAS (n)","Events Control (n)",
    "Readmission Rate ERAS (%)","Readmission Rate Control (%)",
    "Readmission Cause Reported (Y/N)","Top Readmission Cause","Unplanned Readmission Only (Y/N)",
    "Effect Measure (MD/RR/OR/HR/RD)","Effect Size","95% CI Lower","95% CI Upper",
    "p-value","Heterogeneity I2 (%)","Tau2","Statistical Method",
    "Imputation Used (Y/N)","Notes","Extracted By","Date Extracted","Verified By"
]
O3_MAP = {
    "Readmission Timeframe (30-day/90-day/1-year/Other)": "readmission_timeframe",
    "Timeframe — Other (specify days)":                    "timeframe_other",
    "Events ERAS (n)":                                     "events_eras",
    "Events Control (n)":                                  "events_control",
    "Readmission Rate ERAS (%)":                           "readmit_rate_eras",
    "Readmission Rate Control (%)":                        "readmit_rate_ctrl",
    "Readmission Cause Reported (Y/N)":                    "cause_reported",
    "Top Readmission Cause":                               "top_cause",
    "Unplanned Readmission Only (Y/N)":                    "unplanned_only",
}

O4_COLS = [
    "Study ID","First Author","Year","Country","WHO Region","Income Group",
    "Surgical Specialty","Specific Procedure","Study Design","N Total","N ERAS","N Control",
    "Adherence Rate (%)","Implementation Type",
    "Mortality Type (In-hospital/30-day/90-day/1-year/Other)",
    "Events ERAS (n)","Events Control (n)",
    "Mortality Rate ERAS (%)","Mortality Rate Control (%)",
    "Cause of Death Reported (Y/N)","Operative vs Non-operative Mortality",
    "Effect Measure (MD/RR/OR/HR/RD)","Effect Size","95% CI Lower","95% CI Upper",
    "p-value","Heterogeneity I2 (%)","Tau2","Statistical Method",
    "Imputation Used (Y/N)","Notes","Extracted By","Date Extracted","Verified By"
]
O4_MAP = {
    "Mortality Type (In-hospital/30-day/90-day/1-year/Other)": "mortality_type",
    "Events ERAS (n)":                       "events_eras",
    "Events Control (n)":                    "events_control",
    "Mortality Rate ERAS (%)":               "mort_rate_eras",
    "Mortality Rate Control (%)":            "mort_rate_ctrl",
    "Cause of Death Reported (Y/N)":         "cause_reported",
    "Operative vs Non-operative Mortality":  "operative_vs_nonop",
}

O5_COLS = [
    "Study ID","First Author","Year","Country","WHO Region","Income Group",
    "Surgical Specialty","Specific Procedure","Study Design","N Total","N ERAS","N Control",
    "Adherence Rate (%)","Implementation Type",
    "Cost Perspective (Hospital/Societal/Payer/Patient)","Cost Components (Direct/Indirect/Both)",
    "Year of Costing","Original Currency",
    "Total Cost — ERAS (original currency)","Total Cost — Control (original currency)",
    "Cost Difference (original currency)","PPP Conversion Rate Used",
    "Total Cost — ERAS (USD 2023 PPP)","Total Cost — Control (USD 2023 PPP)",
    "Cost Difference (USD 2023 PPP)","% Cost Reduction","ICER Reported (Y/N)","ICER Value",
    "Cost per Bed-Day Reported (Y/N)","Sensitivity Analysis Reported (Y/N)",
    "Effect Measure (MD/RR/OR/HR/RD)","Effect Size","95% CI Lower","95% CI Upper",
    "p-value","Heterogeneity I2 (%)","Tau2","Statistical Method",
    "Imputation Used (Y/N)","Notes","Extracted By","Date Extracted","Verified By"
]
O5_MAP = {
    "Cost Perspective (Hospital/Societal/Payer/Patient)": "cost_perspective",
    "Cost Components (Direct/Indirect/Both)":             "cost_components",
    "Year of Costing":                                    "year_of_costing",
    "Original Currency":                                  "original_currency",
    "Total Cost — ERAS (original currency)":              "cost_eras_orig",
    "Total Cost — Control (original currency)":           "cost_ctrl_orig",
    "Cost Difference (original currency)":                "cost_diff_orig",
    "PPP Conversion Rate Used":                           "ppp_rate",
    "Total Cost — ERAS (USD 2023 PPP)":                   "cost_eras_usd",
    "Total Cost — Control (USD 2023 PPP)":                "cost_ctrl_usd",
    "Cost Difference (USD 2023 PPP)":                     "cost_diff_usd",
    "% Cost Reduction":                                   "pct_cost_reduction",
    "ICER Reported (Y/N)":                                "icer_reported",
    "ICER Value":                                         "icer_value",
    "Cost per Bed-Day Reported (Y/N)":                    "cost_per_bedday",
    "Sensitivity Analysis Reported (Y/N)":                "sensitivity_analysis",
}

COMMON_MAP = {
    "Study ID":              "study_id",
    "First Author":          "first_author",
    "Year":                  "year",
    "Country":               "country",
    "WHO Region":            "who_region",
    "Income Group":          "income_group",
    "Surgical Specialty":    "surgical_specialty",
    "Specific Procedure":    "specific_procedure",
    "Study Design":          "study_design",
    "N Total":               "n_total",
    "N ERAS":                "n_eras",
    "N Control":             "n_control",
    "Adherence Rate (%)":    "adherence_rate",
    "Implementation Type":   "implementation_type",
    "Effect Measure (MD/RR/OR/HR/RD)": "effect_measure",
    "Effect Size":           "effect_size",
    "95% CI Lower":          "ci_lower",
    "95% CI Upper":          "ci_upper",
    "p-value":               "p_value",
    "Heterogeneity I2 (%)":  "i2",
    "Tau2":                  "tau2",
    "Statistical Method":    "stat_method",
    "Imputation Used (Y/N)": "imputation_used",
    "Notes":                 "notes",
    "Extracted By":          "extracted_by",
    "Date Extracted":        "date_extracted",
    "Verified By":           "verified_by",
}

ADH_ELEM_MAP = {el_label: el_key for el_key, el_label in ERAS_ELEMENTS}
ADH_ELEM_MAP.update({
    "Other Element 1 (specify in Notes)": "other_element1",
    "Other Element 2 (specify in Notes)": "other_element2",
})

ADH_COLS = [
    "Study ID","First Author","Year","Country","WHO Region","Income Group",
    "Surgical Specialty","Study Design",
    "Total Protocol Elements (n)","Total Implemented Elements (n)",
    "Overall Adherence Rate (%)","Adherence Category (<50% Partial / 50-70% Moderate / >70% Comprehensive)",
    "Adherence Measurement Method (Audit/Checklist/Self-report/Other)",
    "Pre-operative Phase Adherence (%)","Intra-operative Phase Adherence (%)","Post-operative Phase Adherence (%)",
] + [lbl for _, lbl in ERAS_ELEMENTS] + [
    "LOS Difference Associated (days)","Complication OR/RR Associated",
    "p-value for Adherence-Outcome Correlation",
    "Adherence Threshold Reported for Clinical Benefit (Y/N)",
    "Reported Threshold Value (%)","Dose-Response Analysis Reported (Y/N)","Notes"
]
ADH_MAP = {
    "Total Protocol Elements (n)":     "total_elements",
    "Total Implemented Elements (n)":  "implemented_elements",
    "Overall Adherence Rate (%)":      "overall_adherence_rate",
    "Adherence Category (<50% Partial / 50-70% Moderate / >70% Comprehensive)": "adherence_category",
    "Adherence Measurement Method (Audit/Checklist/Self-report/Other)": "adherence_method",
    "Pre-operative Phase Adherence (%)": "preop_phase_adh",
    "Intra-operative Phase Adherence (%)": "intraop_phase_adh",
    "Post-operative Phase Adherence (%)": "postop_phase_adh",
    "LOS Difference Associated (days)": "los_diff_assoc",
    "Complication OR/RR Associated":    "comp_or_rr",
    "p-value for Adherence-Outcome Correlation": "p_assoc",
    "Adherence Threshold Reported for Clinical Benefit (Y/N)": "threshold_reported",
    "Reported Threshold Value (%)":    "threshold_value",
    "Dose-Response Analysis Reported (Y/N)": "dose_response",
    "Notes":                           "notes",
    **{lbl: key for key, lbl in ERAS_ELEMENTS},
}

GEO_COLS = [
    "Study ID","First Author","Year","Country","WHO Region","World Bank Income Group",
    "Sub-region (East Africa / South Asia / etc.)","Urban/Rural Classification",
    "Hospital Level","National GDP per Capita (USD)","Health Expenditure per Capita (USD)",
    "Surgical Workforce Density (per 100k)","Anaesthetist Density (per 100k)","ICU Beds per 100k",
    "Surgical Specialty","Overall ERAS Adherence Rate (%)",
    "LOS — ERAS (days)","LOS — Control (days)","LOS Difference (days)",
    "Complication Rate — ERAS (%)","Complication Rate — Control (%)","RR/OR Complications",
    "Readmission Rate — ERAS (%)","Mortality Rate — ERAS (%)",
    "Cost Savings Reported (Y/N)","Cost Savings (USD PPP)",
    "Implementation Barriers Reported (Y/N)",
    "Barrier — Infrastructure (Y/N)","Barrier — Staff Training (Y/N)",
    "Barrier — Supply Chain (Y/N)","Barrier — Cultural/Patient (Y/N)",
    "Barrier — Financial (Y/N)","Barrier — Other (specify)",
    "Enablers Reported (Y/N)","Enabler — Leadership (Y/N)",
    "Enabler — Training Programme (Y/N)","Enabler — Protocol Adaptation (Y/N)",
    "Dose-Response Analysis Performed (Y/N)","Dose-Response Coefficient (if reported)","Notes"
]
GEO_MAP = {
    "World Bank Income Group":             "income_group",
    "Sub-region (East Africa / South Asia / etc.)": "subregion",
    "Urban/Rural Classification":          "urban_rural",
    "Hospital Level":                      "hospital_level",
    "National GDP per Capita (USD)":       "gdp_per_capita",
    "Health Expenditure per Capita (USD)": "health_exp_per_cap",
    "Surgical Workforce Density (per 100k)": "surg_workforce_per100k",
    "Anaesthetist Density (per 100k)":     "anaesthetist_per100k",
    "ICU Beds per 100k":                   "icu_beds_per100k",
    "Overall ERAS Adherence Rate (%)":     "overall_adherence_rate",
    "LOS — ERAS (days)":                   "los_eras",
    "LOS — Control (days)":               "los_ctrl",
    "LOS Difference (days)":              "los_diff",
    "Complication Rate — ERAS (%)":       "comp_rate_eras",
    "Complication Rate — Control (%)":    "comp_rate_ctrl",
    "RR/OR Complications":                "rr_or_comp",
    "Readmission Rate — ERAS (%)":        "readmit_rate_eras",
    "Mortality Rate — ERAS (%)":          "mort_rate_eras",
    "Cost Savings Reported (Y/N)":        "cost_savings_reported",
    "Cost Savings (USD PPP)":             "cost_savings_usd",
    "Implementation Barriers Reported (Y/N)": "barriers_reported",
    "Barrier — Infrastructure (Y/N)":     "b_infrastructure",
    "Barrier — Staff Training (Y/N)":     "b_training",
    "Barrier — Supply Chain (Y/N)":       "b_supply",
    "Barrier — Cultural/Patient (Y/N)":   "b_cultural",
    "Barrier — Financial (Y/N)":          "b_financial",
    "Barrier — Other (specify)":          "b_other",
    "Enablers Reported (Y/N)":            "enablers_reported",
    "Enabler — Leadership (Y/N)":         "e_leadership",
    "Enabler — Training Programme (Y/N)": "e_training",
    "Enabler — Protocol Adaptation (Y/N)":"e_protocol",
    "Dose-Response Analysis Performed (Y/N)": "dose_resp_analysis",
    "Dose-Response Coefficient (if reported)": "dose_resp_coef",
    "Notes":                              "notes",
}


def build_rows(records: list, col_list: list, extra_map: dict) -> list[dict]:
    """Flatten records to {col_header: value} dicts."""
    rows = []
    col_map = {**COMMON_MAP, **extra_map}
    for rec in records:
        row = {}
        for col in col_list:
            key = col_map.get(col, None)
            row[col] = rec.get(key, "") if key else ""
        rows.append(row)
    return rows


def build_excel() -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    # Collect all records across studies
    def all_records(outcome):
        out = []
        for si in range(len(ss.studies)):
            out.extend(get_records(si, outcome))
        return out

    sheet_defs = [
        ("O1_LOS",  "O1_LOS — LENGTH OF STAY | ERAS Systematic Review Data Extraction",
         O1_COLS, O1_MAP),
        ("O2_COMP", "O2_COMP — POSTOPERATIVE COMPLICATIONS | ERAS Systematic Review Data Extraction",
         O2_COLS, O2_MAP),
        ("O3_READ", "O3_READ — READMISSION RATES | ERAS Systematic Review Data Extraction",
         O3_COLS, O3_MAP),
        ("O4_MORT", "O4_MORT — MORTALITY | ERAS Systematic Review Data Extraction",
         O4_COLS, O4_MAP),
        ("O5_COST", "O5_COST — COST-EFFECTIVENESS | ERAS Systematic Review Data Extraction",
         O5_COLS, O5_MAP),
        ("ADHERENCE","ADHERENCE — ERAS Protocol Compliance | ERAS Systematic Review Data Extraction",
         ADH_COLS, ADH_MAP),
        ("GEO_DISPARITY","GEO_DISPARITY — Geographical Variation | ERAS Systematic Review Data Extraction",
         GEO_COLS, GEO_MAP),
    ]
    for outcome, title, cols, extra_map in sheet_defs:
        ws = wb.create_sheet(outcome)
        recs = all_records(outcome)
        rows = build_rows(recs, cols, extra_map)
        _write_sheet(ws, cols, rows, title)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Progress indicator ────────────────────────────────────────────────────────
def outcome_status(study_idx: int, outcome: str) -> str:
    recs = get_records(study_idx, outcome)
    if not recs:
        return "⬜"
    return "✅"


# ── MAIN LAYOUT ───────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 🔬 ERAS Data Extractor")
    st.caption("Systematic Review · 7 Outcome Domains")

    # Upload
    uploaded = st.file_uploader(
        "Upload full-text PDFs",
        type="pdf",
        accept_multiple_files=True,
        label_visibility="collapsed",
    )
    if uploaded:
        new_files = [f for f in uploaded if f.name not in ss.uploaded_names]
        if new_files:
            with st.status(f"Processing {len(new_files)} PDF(s)…", expanded=True) as _st:
                for f in new_files:
                    st.write(f"📄 {f.name[:50]}")
                    text = extract_pdf(f.read())
                    ss.studies.append({"name": f.name, "text": text})
                    ss.uploaded_names.add(f.name)
                _st.update(label=f"✅ {len(ss.studies)} study(ies) loaded", state="complete", expanded=False)
            st.rerun()

    if not ss.studies:
        st.info("Upload one or more PDFs to begin.")
        st.stop()

    st.markdown("---")
    st.markdown("**Studies**")
    for i, study in enumerate(ss.studies):
        progress = "".join(outcome_status(i, o) for o in OUTCOMES)
        label = f"{study['name'][:32]}…" if len(study['name']) > 35 else study['name']
        if st.button(f"{label}\n{progress}", key=f"nav_{i}",
                     type="primary" if i == ss.current_study else "secondary",
                     use_container_width=True):
            ss.current_study = i
            st.rerun()

    st.markdown("---")
    # Save / Load JSON
    st.markdown("**Progress**")
    json_data = json.dumps({
        "studies": [s["name"] for s in ss.studies],
        "extractions": {str(k): v for k, v in ss.extractions.items()}
    }, indent=2, default=str)
    st.download_button("⬇ Save progress (JSON)",
                       data=json_data.encode(),
                       file_name="eras_extraction_progress.json",
                       mime="application/json",
                       use_container_width=True)
    load_file = st.file_uploader("↑ Load progress (JSON)", type="json",
                                  label_visibility="collapsed", key="json_load")
    if load_file:
        try:
            loaded = json.load(load_file)
            ss.extractions = {int(k): v for k, v in loaded.get("extractions", {}).items()}
            st.success("Progress loaded!"); st.rerun()
        except Exception as e:
            st.error(f"Load error: {e}")

    st.markdown("---")
    # Excel export
    if st.button("🔄 Generate Excel", use_container_width=True, type="primary"):
        with st.spinner("Building workbook…"):
            ss.excel_bytes = build_excel()
        st.success("Ready ↓")
    if ss.excel_bytes:
        st.download_button("⬇ Download Excel (.xlsx)",
                           data=ss.excel_bytes,
                           file_name=f"ERAS_DataExtraction_{date.today()}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)


# ── Current study ─────────────────────────────────────────────────────────────
idx = ss.current_study
study = ss.studies[idx]

st.markdown(f"### 📄 {study['name']}")
st.caption(f"Study {idx+1} of {len(ss.studies)}  ·  "
           + "  ".join(f"{OUTCOME_ICONS[o]} {outcome_status(idx, o)}"
                       for o in OUTCOMES))

# Search
ss.search_term = st.text_input("🔍 Search in text (highlights in orange)",
                                ss.search_term, placeholder="e.g. length of stay, median, complication…",
                                label_visibility="collapsed")

col_text, col_form = st.columns([5, 5])

with col_text:
    st.markdown("**Full-Text Viewer**")
    viewer_html = build_viewer(study["text"], ss.search_term)
    st.markdown(viewer_html, unsafe_allow_html=True)
    st.caption("Colours: 🟡 LOS  🟠 Complications  🟣 Readmission  🔴 Mortality  🔵 Cost  🟢 Adherence  🔶 Search")

with col_form:
    st.markdown("**Data Extraction**")
    tabs = st.tabs([OUTCOME_ICONS[o] for o in OUTCOMES])

    with tabs[0]:
        st.caption("Length of Stay — supports multiple records (Hospital LOS + ICU LOS as separate rows)")
        _b1, _b2 = st.columns([1, 1])
        if _b1.button("🤖 Smart Extract", key=f"smart_O1_{idx}", type="primary",
                      help="Auto-read article text and pre-fill form — review before saving"):
            _apply_extracted("O1_LOS", idx, smart_extract_all(study["text"], "O1_LOS"))
            st.rerun()
        if _b2.button("➕ New LOS Record", key=f"new_O1_LOS_{idx}"):
            _apply_extracted("O1_LOS", idx, {}); st.rerun()
        form_o1(idx)

    with tabs[1]:
        st.caption("Complications — add one record per complication type (Overall + each specific type)")
        _b1, _b2 = st.columns([1, 1])
        if _b1.button("🤖 Smart Extract", key=f"smart_O2_{idx}", type="primary",
                      help="Auto-read article text and pre-fill form — review before saving"):
            _apply_extracted("O2_COMP", idx, smart_extract_all(study["text"], "O2_COMP"))
            st.rerun()
        if _b2.button("➕ New Complication Record", key=f"new_O2_COMP_{idx}"):
            _apply_extracted("O2_COMP", idx, {}); st.rerun()
        form_o2(idx)

    with tabs[2]:
        st.caption("Readmission — typically one record per study (or one per timeframe)")
        if st.button("🤖 Smart Extract", key=f"smart_O3_{idx}", type="primary",
                     help="Auto-read article text and pre-fill form"):
            _apply_extracted("O3_READ", idx, smart_extract_all(study["text"], "O3_READ"))
            st.rerun()
        form_o3(idx)

    with tabs[3]:
        st.caption("Mortality — add separate records for in-hospital and 30-day if both reported")
        if st.button("🤖 Smart Extract", key=f"smart_O4_{idx}", type="primary",
                     help="Auto-read article text and pre-fill form"):
            _apply_extracted("O4_MORT", idx, smart_extract_all(study["text"], "O4_MORT"))
            st.rerun()
        form_o4(idx)

    with tabs[4]:
        st.caption("Cost-Effectiveness — PPP-adjust all costs to USD 2023 using World Bank rates")
        if st.button("🤖 Smart Extract", key=f"smart_O5_{idx}", type="primary",
                     help="Auto-read article text and pre-fill form"):
            _apply_extracted("O5_COST", idx, smart_extract_all(study["text"], "O5_COST"))
            st.rerun()
        form_o5(idx)

    with tabs[5]:
        st.caption("ERAS Protocol Adherence — element-level compliance (0=Not implemented, 1=Full, 0.5=Partial, or %)")
        if st.button("🤖 Smart Extract", key=f"smart_ADH_{idx}", type="primary",
                     help="Auto-detect ERAS elements and adherence rate from text"):
            _apply_extracted("ADHERENCE", idx, smart_extract_all(study["text"], "ADHERENCE"))
            st.rerun()
        form_adherence(idx)

    with tabs[6]:
        st.caption("Geographical variation — include country-level contextual data from World Bank (study year)")
        if st.button("🤖 Smart Extract", key=f"smart_GEO_{idx}", type="primary",
                     help="Auto-fill country, region, income group from text"):
            _apply_extracted("GEO_DISPARITY", idx, smart_extract_all(study["text"], "GEO_DISPARITY"))
            st.rerun()
        form_geo(idx)
